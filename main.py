import os, json, httpx, asyncio, time, io, zipfile, calendar, csv, re
from uuid import uuid4
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, datetime, timedelta
import dropbox
from dropbox.exceptions import ApiError as DropboxApiError
import gspread
from google.oauth2.service_account import Credentials
from fastapi import FastAPI, Request, UploadFile, File, Form, Query, Header, HTTPException, Response
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse, HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles

app = FastAPI()

DROPBOX_APP_KEY      = os.environ.get("DROPBOX_APP_KEY")
DROPBOX_APP_SECRET   = os.environ.get("DROPBOX_APP_SECRET")
DROPBOX_REFRESH_TOKEN = os.environ.get("DROPBOX_REFRESH_TOKEN")
SHIKONSHOSAI_APP_URL = os.environ.get("SHIKONSHOSAI_APP_URL", "https://shikonshosai-app.onrender.com")
INTERNAL_SECRET      = os.environ.get("INTERNAL_SECRET", "shikonshosai_internal_2024")
SPREADSHEET_ID   = "1Pt4Mvzp11FMWLxwF-iFYxF2iVHVn54qNpaJRhkicDe4"
FREEE_COMPANY_ID = int(os.environ.get("FREEE_COMPANY_ID", "3254695"))
_FREEE_CLIENT_ID     = os.environ.get("FREEE_CLIENT_ID", "")
_FREEE_CLIENT_SECRET = os.environ.get("FREEE_CLIENT_SECRET", "")
_FREEE_REFRESH_TOKEN = os.environ.get("FREEE_REFRESH_TOKEN", "")
# freee 人事労務 API テスト用
_FREEE_HR_CLIENT_ID     = os.environ.get("FREEE_HR_CLIENT_ID", "")
_FREEE_HR_CLIENT_SECRET = os.environ.get("FREEE_HR_CLIENT_SECRET", "")
_FREEE_HR_REDIRECT_URI  = os.environ.get(
    "FREEE_HR_REDIRECT_URI",
    "https://shikonshosai-manual.onrender.com/api/hr/callback",
)
HR_TOKEN_PATH = "/外注先共有/400000_CC/shikonshosai/hr_token.json"
FREEE_HR_API_BASE = "https://api.freee.co.jp/hr/api/v1"
_stored_refresh_token: str = ""
_token_lock = None
FREEE_AUTH_URL  = "https://accounts.secure.freee.co.jp/public_api/authorize"
FREEE_TOKEN_URL = "https://accounts.secure.freee.co.jp/public_api/token"
FREEE_REDIRECT_URI = os.environ.get(
    "FREEE_REDIRECT_URI",
    "https://shikonshosai-manual.onrender.com/auth/callback",
)
MANUALS_PATH  = "/外注先共有/400000_CC/shikonshosai/manuals.json"
NOTICES_PATH  = "/外注先共有/400000_CC/shikonshosai/notices.json"
QA_PATH       = "/外注先共有/400000_CC/shikonshosai/qa.json"
USERS_PATH    = "/外注先共有/400000_CC/shikonshosai/users.json"
IMAGES_BASE   = "/外注先共有/400000_CC/shikonshosai/manual_images"
REPORTS_BASE  = "/外注先共有/400000_CC/shikonshosai/reports"
COMPANIES_PATH         = "/外注先共有/400000_CC/shikonshosai/companies.json"
COMPANY_PROGRESS_BASE  = "/外注先共有/400000_CC/shikonshosai/company_progress/"
COMPANY_MANUALS_BASE   = "/外注先共有/400000_CC/shikonshosai/company_manuals"
COMPANY_SCHEDULES_BASE = "/外注先共有/400000_CC/shikonshosai/company_schedules"
CACHE_TTL_COMPANIES    = 30 * 24 * 60 * 60  # 30日
def get_invoices_path(year_month: str) -> str:
    year = year_month.split("-")[0] if year_month else str(date.today().year)
    return f"/外注先共有/400000_CC/shikonshosai/invoices_{year}.json"

def get_pledges_path(year_month: str) -> str:
    year = year_month.split("-")[0] if year_month else str(date.today().year)
    return f"/外注先共有/400000_CC/shikonshosai/pledges_{year}.json"

_cache: dict = {}
_CACHE_TTL = 60

def _cache_get(key):
    if key in _cache:
        entry = _cache[key]
        if len(entry) == 3:
            data, ts, ttl = entry
        else:
            data, ts = entry
            ttl = _CACHE_TTL
        if time.time() - ts < ttl:
            return data
    return None

def _cache_set(key, data, ttl=None):
    _cache[key] = (data, time.time(), ttl if ttl is not None else _CACHE_TTL)

def _cache_delete(key):
    _cache.pop(key, None)

def _get_dropbox_client():
    return dropbox.Dropbox(
        oauth2_refresh_token=DROPBOX_REFRESH_TOKEN,
        app_key=DROPBOX_APP_KEY,
        app_secret=DROPBOX_APP_SECRET,
    )

async def dropbox_get(path: str):
    try:
        dbx = _get_dropbox_client()
        _, res = dbx.files_download(path)
        return json.loads(res.content)
    except DropboxApiError as e:
        if e.error.is_path() and e.error.get_path().is_not_found():
            return None
        raise

async def dropbox_save(path: str, data: dict):
    dbx = _get_dropbox_client()
    content = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
    dbx.files_upload(content, path, mode=dropbox.files.WriteMode.overwrite, mute=True)

async def dropbox_delete(path: str):
    try:
        dbx = _get_dropbox_client()
        dbx.files_delete_v2(path)
    except DropboxApiError:
        pass

def _get_gspread_client():
    creds_json = os.environ.get("GOOGLE_CREDENTIALS_JSON", "")
    creds_dict = json.loads(creds_json)
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    return gspread.authorize(creds)

def _get_spreadsheet():
    gc = _get_gspread_client()
    return gc.open_by_key(SPREADSHEET_ID)

_freee_access_token: str = ""
_freee_token_expires_at: float = 0.0

async def _update_render_env(new_refresh_token: str):
    """Renderの環境変数FREEE_REFRESH_TOKENを最新値に更新する"""
    api_key    = os.environ.get("RENDER_API_KEY", "")
    service_id = os.environ.get("RENDER_SERVICE_ID", "")
    if not api_key or not service_id:
        return
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            await client.put(
                f"https://api.render.com/v1/services/{service_id}/env-vars/FREEE_REFRESH_TOKEN",
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                },
                json={"value": new_refresh_token},
            )
    except Exception:
        pass  # 失敗してもトークン自体は更新済みなので無視

async def _get_freee_token() -> str:
    """
    トークン取得：_stored_refresh_token（/auth/callback で更新）を優先し、
    なければ環境変数 FREEE_REFRESH_TOKEN を使ってアクセストークンを更新する。
    新しい refresh_token が返ってきたら _stored_refresh_token に保存する。
    """
    global _freee_access_token, _freee_token_expires_at, _stored_refresh_token, _token_lock
    if _freee_access_token and time.time() < _freee_token_expires_at - 60:
        return _freee_access_token
    if _token_lock is None:
        _token_lock = asyncio.Lock()
    async with _token_lock:
        if _freee_access_token and time.time() < _freee_token_expires_at - 60:
            return _freee_access_token
        refresh_token = _stored_refresh_token or _FREEE_REFRESH_TOKEN
        async with httpx.AsyncClient() as client:
            r = await client.post(
                FREEE_TOKEN_URL,
                data={
                    "grant_type": "refresh_token",
                    "client_id": _FREEE_CLIENT_ID,
                    "client_secret": _FREEE_CLIENT_SECRET,
                    "refresh_token": refresh_token,
                }
            )
            r.raise_for_status()
            td = r.json()
            _freee_access_token = td["access_token"]
            _freee_token_expires_at = time.time() + td.get("expires_in", 3600)
            if "refresh_token" in td:
                _stored_refresh_token = td["refresh_token"]
                asyncio.create_task(_update_render_env(td["refresh_token"]))
            return _freee_access_token

@app.get("/auth/login")
async def freee_auth_login():
    if not _FREEE_CLIENT_ID:
        raise HTTPException(status_code=500, detail="FREEE_CLIENT_ID 未設定")
    url = (
        f"{FREEE_AUTH_URL}"
        f"?client_id={_FREEE_CLIENT_ID}"
        f"&redirect_uri={FREEE_REDIRECT_URI}"
        f"&response_type=code"
    )
    return RedirectResponse(url)

@app.get("/auth/callback")
async def freee_auth_callback(code: str = Query(...)):
    global _stored_refresh_token, _freee_access_token, _freee_token_expires_at
    if not _FREEE_CLIENT_ID or not _FREEE_CLIENT_SECRET:
        raise HTTPException(status_code=500, detail="FREEE_CLIENT_ID/SECRET 未設定")
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            FREEE_TOKEN_URL,
            data={
                "grant_type": "authorization_code",
                "client_id": _FREEE_CLIENT_ID,
                "client_secret": _FREEE_CLIENT_SECRET,
                "code": code,
                "redirect_uri": FREEE_REDIRECT_URI,
            }
        )
        resp.raise_for_status()
        tokens = resp.json()
    _stored_refresh_token = tokens["refresh_token"]
    _freee_access_token = tokens["access_token"]
    _freee_token_expires_at = time.time() + tokens.get("expires_in", 3600)
    return HTMLResponse(
        "<html><body style='font-family:sans-serif;padding:40px;text-align:center;'>"
        "<h2>freee 認証完了</h2>"
        "<p>このタブを閉じてください。</p>"
        "</body></html>"
    )

# ========================================
# freee 人事労務 API テスト用エンドポイント
# ========================================

async def _hr_require_admin(user_id: str):
    if not user_id:
        raise HTTPException(status_code=401, detail="ログインが必要です")
    users_data = await dropbox_get(USERS_PATH) or {"users": []}
    user = next((u for u in users_data.get("users", []) if u.get("id") == user_id), None)
    if not user or user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="admin only")
    return user


async def _hr_get_token() -> dict:
    """hr_token.json を読み込む。無ければ None。"""
    return await dropbox_get(HR_TOKEN_PATH)


async def _hr_save_token(access_token: str, refresh_token: str, expires_in: int):
    expires_at = int(time.time()) + int(expires_in or 3600)
    await dropbox_save(HR_TOKEN_PATH, {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "expires_at": expires_at,
    })


async def _hr_refresh_access_token(refresh_token: str) -> dict:
    """refresh_token を使ってアクセストークンを再取得し hr_token.json を更新。"""
    if not _FREEE_HR_CLIENT_ID or not _FREEE_HR_CLIENT_SECRET:
        raise HTTPException(status_code=500, detail="FREEE_HR_CLIENT_ID/SECRET 未設定")
    async with httpx.AsyncClient() as client:
        r = await client.post(FREEE_TOKEN_URL, data={
            "grant_type": "refresh_token",
            "client_id": _FREEE_HR_CLIENT_ID,
            "client_secret": _FREEE_HR_CLIENT_SECRET,
            "refresh_token": refresh_token,
        })
        if r.status_code != 200:
            raise HTTPException(status_code=502, detail=f"freee refresh失敗: {r.text}")
        td = r.json()
    await _hr_save_token(td["access_token"], td.get("refresh_token", refresh_token), td.get("expires_in", 3600))
    return await _hr_get_token()


def _extract_employees_list(emp_data):
    """freee /employees のレスポンス構造ゆれを吸収して list を返す。
    ありうる形：[...] / {"employees": [...]} / {"data": [...]} / {"results": [...]}"""
    if isinstance(emp_data, list):
        return emp_data
    if isinstance(emp_data, dict):
        for key in ("employees", "data", "results", "items"):
            v = emp_data.get(key)
            if isinstance(v, list):
                return v
    return []


def _find_employee_by_num(emp_list, target_num):
    """freee 人事労務 /employees レスポンスから num が一致する従業員を返す。
    値の型ゆれ（"018" / "18" / 18）を吸収するため、文字列一致と整数一致の両方で照合する。"""
    target_str = str(target_num).strip()
    try:
        target_int = int(target_str)
    except (TypeError, ValueError):
        target_int = None
    print(f"[hr-test] _find_employee_by_num target_str={target_str!r} target_int={target_int!r} count={len(emp_list) if isinstance(emp_list, list) else 'NOT_LIST'}")
    if not isinstance(emp_list, list):
        return None
    for emp in emp_list:
        if not isinstance(emp, dict):
            continue
        raw = emp.get("num")
        if raw is None or raw == "":
            continue
        s = str(raw).strip()
        print(f"[hr-test]   compare emp_id={emp.get('id')!r} num_raw={raw!r} num_str={s!r}")
        if s == target_str:
            return emp
        if target_int is not None:
            try:
                if int(s) == target_int:
                    return emp
            except (TypeError, ValueError):
                pass
    return None


async def _hr_get_valid_access_token() -> str:
    """hr_token.json から有効なアクセストークンを取得。期限切れなら自動リフレッシュ。"""
    tok = await _hr_get_token()
    if not tok:
        raise HTTPException(status_code=401, detail="hr_token 未保存。/api/hr/auth から認証してください")
    if int(tok.get("expires_at", 0)) - 60 < int(time.time()):
        if not tok.get("refresh_token"):
            raise HTTPException(status_code=401, detail="refresh_token なし。再認証してください")
        tok = await _hr_refresh_access_token(tok["refresh_token"])
    return tok["access_token"]


ATTENDANCE_BASE = "/外注先共有/400000_CC/shikonshosai/attendance"
PAYROLL_MONTHLY_HOURS = 166      # 月平均所定時間（固定）
PAYROLL_TEIGAKU_ZANGYO = 55000   # 定額残業手当（固定）
PAYROLL_KIHON_KOTEI = 289000     # 基本給254,000 + 役職手当20,000 + 期日手当15,000


def _attendance_path(employee_id: int, year_month: str) -> str:
    return f"{ATTENDANCE_BASE}/{employee_id}_{year_month}.json"


def _to_float(s):
    try:
        return float(str(s).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _time_only(s):
    if not s:
        return ""
    m = re.search(r"(\d{1,2}):(\d{2})", str(s))
    return f"{int(m.group(1)):02d}:{m.group(2)}" if m else ""


def _calc_estimated_zangyo_h(clock_in: str, clock_out: str) -> float:
    """出退勤時刻から推定残業時間を算出。
    - 9時より前の出勤は9時扱い
    - 実労働 = 退勤 - max(出勤, 9:00) - 休憩1h
    - 残業 = 実労働 - 所定8h（マイナスは0）"""
    if not clock_in or not clock_out:
        return 0.0
    try:
        ci = datetime.strptime(clock_in, "%H:%M")
        co = datetime.strptime(clock_out, "%H:%M")
    except ValueError:
        return 0.0
    nine = ci.replace(hour=9, minute=0, second=0)
    start = ci if ci > nine else nine
    diff_h = (co - start).total_seconds() / 3600 - 1.0
    return max(0.0, diff_h - 8.0)


def _parse_tot_csv(data: bytes):
    """TOT 勤怠 CSV（Shift-JIS）をパースして year_month と records を返す。
    集計開始日が >=21 日なら翌月を year_month とする（例: 2026-01-21 → 2026-02）。"""
    text = data.decode("cp932", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [r for r in reader]
    # ヘッダー行を検出（"日時" を含む行。互換のため "日付" / "年月日" も許容）
    header_idx = None
    for i, row in enumerate(rows):
        if any(("日時" in (c or "")) or ("日付" in (c or "")) or ("年月日" in (c or "")) for c in row):
            header_idx = i
            break
    if header_idx is None:
        raise HTTPException(400, "CSV のヘッダー行が見つかりません（『日時』を含む列が必要）")
    headers = rows[header_idx]

    def col_index(*keys):
        for k in keys:
            for i, h in enumerate(headers):
                if k in (h or ""):
                    return i
        return None

    idx_date = col_index("日時", "日付", "年月日")
    idx_kind = col_index("勤務日種別", "種別")
    idx_kyuka = col_index("休暇")
    idx_note = col_index("備考")
    idx_in = col_index("出勤")
    idx_out = col_index("退勤")
    idx_sched = col_index("所定")
    idx_overtime = col_index("残業")
    idx_total = col_index("労働")

    def get(row, idx):
        if idx is None or idx >= len(row):
            return ""
        return (row[idx] or "").strip()

    records = []
    for row in rows[header_idx + 1:]:
        if not row or all(not (c or "").strip() for c in row):
            continue
        date_raw = get(row, idx_date)
        m = re.match(r"(\d{4})/(\d{1,2})/(\d{1,2})", date_raw)
        if not m:
            continue
        date_iso = f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        note = get(row, idx_note)
        records.append({
            "date": date_iso,
            "day_kind": get(row, idx_kind),
            "leave_type": get(row, idx_kyuka),
            "note": note,
            "is_telework": "在宅" in note,
            "clock_in": _time_only(get(row, idx_in)),
            "clock_out": _time_only(get(row, idx_out)),
            "scheduled_h": _to_float(get(row, idx_sched)),
            "overtime_h": _to_float(get(row, idx_overtime)),
            "total_h": _to_float(get(row, idx_total)),
        })
    if not records:
        raise HTTPException(400, "勤怠データが1件も取得できませんでした")
    # year_month: 開始日 >=21 なら翌月
    y, mo, d = map(int, records[0]["date"].split("-"))
    if d >= 21:
        if mo == 12:
            y += 1; mo = 1
        else:
            mo += 1
    year_month = f"{y}-{mo:02d}"
    return year_month, records


def _attendance_summary(records):
    work_days = 0
    total_overtime_h = 0.0
    telework_days = 0
    for r in records:
        if not isinstance(r, dict):
            continue
        if r.get("overtime_h") is not None or r.get("scheduled_h") or r.get("clock_in"):
            work_days += 1
        if r.get("overtime_h"):
            total_overtime_h += float(r["overtime_h"] or 0)
        if r.get("is_telework"):
            telework_days += 1
    return {
        "work_days": work_days,
        "total_overtime_h": round(total_overtime_h, 2),
        "telework_days": telework_days,
    }


@app.post("/api/hr/attendance/upload")
async def hr_attendance_upload(
    file: UploadFile = File(...),
    employee_id: int = Form(...),
    user_id: str = Header(None, alias="user-id"),
):
    """TOT CSV をアップロード→パース→Dropbox保存"""
    await _hr_require_admin(user_id)
    raw = await file.read()
    year_month, records = _parse_tot_csv(raw)
    payload = {
        "year_month": year_month,
        "employee_id": int(employee_id),
        "records": records,
        "summary": _attendance_summary(records),
        "saved_at": _now_iso(),
        "saved_by": user_id,
        "source": "tot_csv",
    }
    await dropbox_save(_attendance_path(int(employee_id), year_month), payload)
    return payload


@app.get("/api/hr/attendance")
async def hr_get_attendance(
    user_id: str = Header(None, alias="user-id"),
    year_month: str = Query(...),
    employee_id: int = Query(...),
):
    """Dropbox に保存済みの勤怠データを返す。無ければ 404。"""
    await _hr_require_admin(user_id)
    saved = await dropbox_get(_attendance_path(int(employee_id), year_month))
    if not saved:
        raise HTTPException(404, "保存済み勤怠データがありません。CSV をアップロードしてください")
    # summary は最新で再計算
    saved["summary"] = _attendance_summary(saved.get("records") or [])
    return saved


@app.post("/api/hr/attendance")
async def hr_save_attendance(
    request: Request,
    user_id: str = Header(None, alias="user-id"),
):
    """編集済み勤怠データを Dropbox に上書き保存"""
    await _hr_require_admin(user_id)
    body = await request.json()
    employee_id = body.get("employee_id")
    year_month = body.get("year_month")
    if not employee_id or not year_month:
        raise HTTPException(400, "employee_id / year_month が必要です")
    records = body.get("records") or []
    payload = {
        "year_month": year_month,
        "employee_id": int(employee_id),
        "records": records,
        "summary": _attendance_summary(records),
        "saved_at": _now_iso(),
        "saved_by": user_id,
        "source": body.get("source") or "manual_edit",
    }
    await dropbox_save(_attendance_path(int(employee_id), year_month), payload)
    return {"status": "ok", "summary": payload["summary"]}


@app.post("/api/hr/payroll")
async def hr_payroll(
    request: Request,
    user_id: str = Header(None, alias="user-id"),
):
    """保存済み勤怠データから残業・在宅手当を計算して返す（freee送信なし）。
    基礎給与は KIHON_KOTEI (289,000) + 在宅手当（5000×在宅日/勤務日）で自動算出。
    残業時間は出退勤時刻から推定再計算した値を採用し、CSV残業合計と差額も返す。"""
    await _hr_require_admin(user_id)
    body = await request.json()
    employee_id = int(body.get("employee_id") or 0)
    year_month = body.get("year_month") or ""
    if not employee_id or not year_month:
        raise HTTPException(400, "employee_id / year_month が必要です")
    saved = await dropbox_get(_attendance_path(employee_id, year_month))
    if not saved:
        raise HTTPException(404, "保存済み勤怠データがありません。CSV をアップロードしてください")
    records = saved.get("records") or []

    # 集計
    csv_zangyo_h = 0.0
    jitsu_zangyo_h = 0.0  # 推定ロジック計算値
    telework_days = 0
    work_days = 0
    for r in records:
        if not isinstance(r, dict):
            continue
        # CSV の残業時間列合計
        if r.get("overtime_h"):
            csv_zangyo_h += float(r["overtime_h"] or 0)
        # 推定残業時間（出退勤から再計算）
        jitsu_zangyo_h += _calc_estimated_zangyo_h(r.get("clock_in") or "", r.get("clock_out") or "")
        # 在宅日数・勤務日数
        if r.get("is_telework"):
            telework_days += 1
        if r.get("overtime_h") is not None or r.get("scheduled_h") or r.get("clock_in"):
            work_days += 1

    # 在宅手当（仮置き）
    zaitaku_teate = int(round(5000 * telework_days / work_days)) if work_days > 0 else 0
    # 基礎給与 = 固定 + 在宅手当
    kiso_kyuyo = PAYROLL_KIHON_KOTEI + zaitaku_teate
    jikyu = kiso_kyuyo / PAYROLL_MONTHLY_HOURS
    jikyu_125 = jikyu * 1.25
    minashi_zangyo_h = PAYROLL_TEIGAKU_ZANGYO / jikyu_125 if jikyu_125 > 0 else 0

    # 超過残業（推定ロジックベース）
    if jitsu_zangyo_h > minashi_zangyo_h:
        choka_h = jitsu_zangyo_h - minashi_zangyo_h
        choka_kin = int(round(jikyu_125 * choka_h))
    else:
        choka_h = 0.0
        choka_kin = 0

    return {
        "year_month": year_month,
        "employee_id": employee_id,
        "kihon_kotei": PAYROLL_KIHON_KOTEI,
        "zaitaku_teate": zaitaku_teate,
        "kiso_kyuyo": kiso_kyuyo,
        "jikyu": round(jikyu, 2),
        "jikyu_125": round(jikyu_125, 2),
        "minashi_zangyo_h": round(minashi_zangyo_h, 2),
        "jitsu_zangyo_h": round(jitsu_zangyo_h, 2),
        "csv_zangyo_h": round(csv_zangyo_h, 2),
        "zangyo_sa_h": round(jitsu_zangyo_h - csv_zangyo_h, 2),
        "choka_h": round(choka_h, 2),
        "choka_kin": choka_kin,
        "telework_days": telework_days,
        "work_days": work_days,
        "note": "在宅手当は廃止予定・仮置き。残業時間は出退勤からの推定値（社労士確認待ち）。",
    }


@app.post("/api/hr/payroll/submit")
async def hr_payroll_submit(
    request: Request,
    user_id: str = Header(None, alias="user-id"),
):
    """freee 人事労務の給与明細に手当を書き込むテスト用エンドポイント。
    実装は仕様未確定のため PUT /salaries/employee_pay_statements/{year}/{month} を試行し、
    レスポンスをそのまま返す（運用前に必ず人事労務APIドキュメントで正式な
    エンドポイントとペイロード形式を確認すること）。"""
    await _hr_require_admin(user_id)
    body = await request.json()
    employee_id = int(body.get("employee_id") or 0)
    year_month = body.get("year_month") or ""
    choka_kin = int(body.get("choka_kin") or 0)
    zaitaku_teate = int(body.get("zaitaku_teate") or 0)
    if not employee_id or not year_month:
        raise HTTPException(400, "employee_id / year_month が必要です")
    y, mo = year_month.split("-")
    y = int(y); mo = int(mo)

    access_token = await _hr_get_valid_access_token()
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
    }
    async with httpx.AsyncClient(timeout=30) as client:
        me_resp = await client.get(f"{FREEE_HR_API_BASE}/users/me", headers=headers)
        if me_resp.status_code != 200:
            raise HTTPException(502, f"users/me 失敗: {me_resp.text}")
        companies = (me_resp.json() or {}).get("companies", [])
        company_id = companies[0]["id"] if companies else None
        if not company_id:
            raise HTTPException(400, "company_id が取得できませんでした")

        url = f"{FREEE_HR_API_BASE}/salaries/employee_pay_statements/{y}/{mo:02d}"
        payload = {
            "company_id": company_id,
            "employee_id": employee_id,
            "allowances": [
                {"name": "超過残業手当", "amount": choka_kin},
                {"name": "在宅手当", "amount": zaitaku_teate},
            ],
        }
        print(f"[hr-payroll-submit] PUT {url} payload={payload}")
        resp = await client.put(url, headers=headers, json=payload)
        try:
            resp_body = resp.json()
        except Exception:
            resp_body = {"raw": resp.text}
        print(f"[hr-payroll-submit] status={resp.status_code} body={str(resp_body)[:300]}")

    return JSONResponse({
        "request": payload,
        "freee_status": resp.status_code,
        "freee_response": resp_body,
        "note": "freee人事労務 給与明細APIの正式仕様確認後にエンドポイント／ペイロードを調整してください",
    })


@app.get("/api/manuals")
async def get_manuals():
    cached = _cache_get("manuals")
    if cached is not None:
        return cached
    data = await dropbox_get(MANUALS_PATH) or {"categories": []}
    _cache_set("manuals", data)
    return data

@app.post("/api/manuals")
async def save_manuals(request: Request):
    data = await request.json()
    await dropbox_save(MANUALS_PATH, data)
    _cache_delete("manuals")
    return {"ok": True}

@app.get("/api/notices")
async def get_notices():
    cached = _cache_get("notices")
    if cached is not None:
        return cached
    data = await dropbox_get(NOTICES_PATH) or {"notices": []}
    _cache_set("notices", data)
    return data

@app.post("/api/notices")
async def save_notices(request: Request):
    data = await request.json()
    await dropbox_save(NOTICES_PATH, data)
    _cache_delete("notices")
    return {"ok": True}

@app.get("/api/qa")
async def get_qa():
    cached = _cache_get("qa")
    if cached is not None:
        return cached
    data = await dropbox_get(QA_PATH) or {"questions": []}
    _cache_set("qa", data)
    return data

@app.post("/api/qa")
async def save_qa(request: Request):
    data = await request.json()
    await dropbox_save(QA_PATH, data)
    _cache_delete("qa")
    return {"ok": True}

@app.post("/api/manuals/upload_image")
async def upload_image(manual_id: str = Form(...), file: UploadFile = File(...)):
    dbx = _get_dropbox_client()
    content = await file.read()
    path = f"{IMAGES_BASE}/{manual_id}/{file.filename}"
    dbx.files_upload(content, path, mode=dropbox.files.WriteMode.overwrite, mute=True)
    link = dbx.files_get_temporary_link(path)
    return {"path": path, "url": link.link}

@app.get("/api/manuals/image_url")
async def image_url(path: str = Query(...)):
    dbx = _get_dropbox_client()
    link = dbx.files_get_temporary_link(path)
    return {"url": link.link}

@app.delete("/api/manuals/delete_image")
async def delete_image(request: Request):
    body = await request.json()
    dbx = _get_dropbox_client()
    dbx.files_delete_v2(body["path"])
    return {"ok": True}

@app.get("/api/reports/all/{year_month}")
async def get_all_reports(year_month: str, user_id: str = Header(None)):
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    current_user = next((u for u in users_data.get("users", []) if u.get("id") == user_id), None)
    if not current_user:
        raise HTTPException(status_code=401)
    role = current_user.get("role", "staff")
    if role == "admin":
        target_users = users_data["users"]
    elif role == "leader":
        my_group = current_user.get("group", "")
        target_users = [u for u in users_data["users"] if u.get("group") == my_group]
    else:
        raise HTTPException(status_code=403)

    async def fetch_report(user):
        path = f"{REPORTS_BASE}/{user['id']}_{year_month}.json"
        data = await dropbox_get(path)
        return user["id"], user.get("name", user["id"]), (data or {}).get("entries", [])

    results_list = await asyncio.gather(*[fetch_report(u) for u in target_users])
    results = {uid: {"user_name": uname, "entries": entries} for uid, uname, entries in results_list if entries}
    return results

@app.get("/api/reports/excel/{year_month}")
async def download_reports_excel(
    year_month: str,
    user_id: str = Query(None)
):
    if not user_id:
        raise HTTPException(status_code=401)
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    current_user = next((u for u in users_data.get("users", []) if u.get("id") == user_id), None)
    if not current_user or current_user.get("role", "staff") == "staff":
        raise HTTPException(status_code=403)

    if current_user["role"] == "admin":
        target_users = users_data["users"]
    else:
        my_group = current_user.get("group", "")
        target_users = [u for u in users_data["users"] if u.get("group") == my_group]

    async def fetch_report(user):
        path = f"{REPORTS_BASE}/{user['id']}_{year_month}.json"
        try:
            data = await dropbox_get(path)
            return user.get("name", user["id"]), (data or {}).get("entries", [])
        except Exception:
            return user.get("name", user["id"]), []

    results = await asyncio.gather(*[fetch_report(u) for u in target_users])

    # 会社×スタッフのクロス集計
    companies: dict = {}
    for uname, entries in results:
        for e in entries:
            company = e.get("company_name") or "（会社名なし）"
            if company not in companies:
                companies[company] = {}
            companies[company][uname] = companies[company].get(uname, 0) + e.get("hours", 0)

    staff_names = [r[0] for r in results]
    sorted_companies = sorted(
        companies.items(),
        key=lambda kv: sum(kv[1].values()),
        reverse=True
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year_month}集計"

    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")

    header = ["会社名"] + staff_names + ["合計"]
    for col, val in enumerate(header, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, (company, staff_hours) in enumerate(sorted_companies, 2):
        ws.cell(row=row_idx, column=1, value=company)
        row_total = 0.0
        for col_idx, sname in enumerate(staff_names, 2):
            h = round(staff_hours.get(sname, 0) * 100) / 100
            if h > 0:
                ws.cell(row=row_idx, column=col_idx, value=h)
                row_total += h
        ws.cell(row=row_idx, column=len(staff_names) + 2, value=round(row_total * 100) / 100)

    total_row = len(sorted_companies) + 2
    ws.cell(row=total_row, column=1, value="合計").font = Font(bold=True)
    grand_total = 0.0
    for col_idx, sname in enumerate(staff_names, 2):
        col_total = round(sum(v.get(sname, 0) for v in companies.values()) * 100) / 100
        if col_total > 0:
            ws.cell(row=total_row, column=col_idx, value=col_total).font = Font(bold=True)
        grand_total += col_total
    ws.cell(row=total_row, column=len(staff_names) + 2, value=round(grand_total * 100) / 100).font = Font(bold=True)

    ws.column_dimensions["A"].width = 25
    from openpyxl.utils import get_column_letter
    for i in range(2, len(staff_names) + 3):
        ws.column_dimensions[get_column_letter(i)].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    filename = quote(f"グループ集計_{year_month}.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

@app.get("/api/reports/excel_amount/{year_month}")
async def download_reports_excel_amount(
    year_month: str,
    user_id: str = Query(None)
):
    if not user_id:
        raise HTTPException(status_code=401)
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    current_user = next((u for u in users_data.get("users", []) if u.get("id") == user_id), None)
    if not current_user or current_user.get("role", "staff") == "staff":
        raise HTTPException(status_code=403)

    if current_user["role"] == "admin":
        target_users = users_data["users"]
    else:
        my_group = current_user.get("group", "")
        target_users = [u for u in users_data["users"] if u.get("group") == my_group]

    def _to_rate(v) -> float:
        try:
            return float(v) if v not in (None, "") else 0.0
        except (TypeError, ValueError):
            return 0.0

    rate_by_name = {
        u.get("name", u["id"]): _to_rate(u.get("hourly_rate"))
        for u in target_users
    }

    async def fetch_report(user):
        path = f"{REPORTS_BASE}/{user['id']}_{year_month}.json"
        try:
            data = await dropbox_get(path)
            return user.get("name", user["id"]), (data or {}).get("entries", [])
        except Exception:
            return user.get("name", user["id"]), []

    results = await asyncio.gather(*[fetch_report(u) for u in target_users])

    # 会社×スタッフのクロス集計（金額 = 時間 × 時給）
    companies: dict = {}
    for uname, entries in results:
        rate = rate_by_name.get(uname, 0.0)
        for e in entries:
            company = e.get("company_name") or "（会社名なし）"
            if company not in companies:
                companies[company] = {}
            amount = round((e.get("hours", 0) or 0) * rate)
            companies[company][uname] = companies[company].get(uname, 0) + amount

    staff_names = [r[0] for r in results]
    sorted_companies = sorted(
        companies.items(),
        key=lambda kv: sum(kv[1].values()),
        reverse=True
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "金額集計"

    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")

    header = ["会社名"] + staff_names + ["合計"]
    for col, val in enumerate(header, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, (company, staff_amounts) in enumerate(sorted_companies, 2):
        ws.cell(row=row_idx, column=1, value=company)
        row_total = 0
        for col_idx, sname in enumerate(staff_names, 2):
            a = int(staff_amounts.get(sname, 0))
            if a > 0:
                cell = ws.cell(row=row_idx, column=col_idx, value=a)
                cell.number_format = "#,##0"
                row_total += a
        total_cell = ws.cell(row=row_idx, column=len(staff_names) + 2, value=row_total)
        total_cell.number_format = "#,##0"

    total_row = len(sorted_companies) + 2
    ws.cell(row=total_row, column=1, value="合計").font = Font(bold=True)
    grand_total = 0
    for col_idx, sname in enumerate(staff_names, 2):
        col_total = int(sum(v.get(sname, 0) for v in companies.values()))
        if col_total > 0:
            c = ws.cell(row=total_row, column=col_idx, value=col_total)
            c.font = Font(bold=True)
            c.number_format = "#,##0"
        grand_total += col_total
    gc_cell = ws.cell(row=total_row, column=len(staff_names) + 2, value=grand_total)
    gc_cell.font = Font(bold=True)
    gc_cell.number_format = "#,##0"

    ws.column_dimensions["A"].width = 25
    from openpyxl.utils import get_column_letter
    for i in range(2, len(staff_names) + 3):
        ws.column_dimensions[get_column_letter(i)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    filename = quote(f"グループ金額集計_{year_month}.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

@app.get("/api/reports/{user_id}/{year_month}")
async def get_report(user_id: str, year_month: str):
    path = f"{REPORTS_BASE}/{user_id}_{year_month}.json"
    data = await dropbox_get(path)
    return data if data else {"entries": []}

@app.post("/api/reports/{user_id}/{year_month}")
async def save_report(user_id: str, year_month: str, request: Request):
    invoices_data = await dropbox_get(get_invoices_path(year_month)) or {"invoices": []}
    approved = any(
        inv for inv in invoices_data.get("invoices", [])
        if inv.get("user_id") == user_id
        and inv.get("year_month") == year_month
        and inv.get("status") == "approved"
    )
    if approved:
        raise HTTPException(status_code=403, detail="承認済みの請求書があるため日報を変更できません")
    data = await request.json()
    path = f"{REPORTS_BASE}/{user_id}_{year_month}.json"
    await dropbox_save(path, data)
    return {"ok": True}

@app.on_event("startup")
async def startup_event():
    data = await dropbox_get(USERS_PATH)
    if data is None:
        await dropbox_save(USERS_PATH, {
            "password": "shikonshosai",
            "users": [{"id": "u1", "name": "勝野弘志", "email": "hkcpa416@gmail.com", "role": "admin", "group": "", "photo": ""}]
        })
    else:
        updated = False
        for u in data.get("users", []):
            if "group" not in u:
                u["group"] = ""
                updated = True
        if updated:
            await dropbox_save(USERS_PATH, data)

@app.post("/api/users/profile")
async def update_profile(request: Request):
    body = await request.json()
    user_id = body.get("user_id")
    if not user_id:
        raise HTTPException(status_code=401)
    allowed_fields = [
        "photo", "personal_email", "phone", "postal_code", "address",
        "bank_name", "bank_name_kana", "bank_code", "bank_branch", "branch_kana", "branch_code",
        "bank_type", "bank_number",
        "bank_holder", "invoice_number", "hourly_rate"
    ]
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    updated_user = None
    for user in users_data.get("users", []):
        if user.get("id") == user_id:
            for field in allowed_fields:
                if field in body:
                    user[field] = body[field]
            updated_user = user
            break
    await dropbox_save(USERS_PATH, users_data)
    _cache_delete("users")
    if updated_user:
        asyncio.create_task(_sync_freee_partner(updated_user))
    return {"ok": True}

@app.get("/api/users")
async def get_users():
    cached = _cache_get("users")
    if cached is not None:
        return cached
    data = await dropbox_get(USERS_PATH)
    if data is None:
        return {"users": []}
    result = {"users": [{k: v for k, v in u.items() if k != "individual_password"} for u in data.get("users", [])]}
    _cache_set("users", result)
    return result

@app.post("/api/users")
async def save_users(request: Request):
    data = await request.json()
    # GET /api/users はindividual_passwordを除外して返すため、既存データから引き継ぐ
    # また共通パスワード（data["password"]）はクライアントから書き換えさせない
    existing = await dropbox_get(USERS_PATH)
    deleted_user_ids: set = set()
    if existing:
        data["password"] = existing.get("password", data.get("password", ""))
        existing_pw = {u["id"]: u.get("individual_password", "") for u in existing.get("users", [])}
        for user in data.get("users", []):
            uid = user.get("id", "")
            if uid in existing_pw and existing_pw[uid]:
                user["individual_password"] = existing_pw[uid]
        new_ids = {u.get("id") for u in data.get("users", [])}
        deleted_user_ids = {uid for uid in existing_pw.keys() if uid not in new_ids}
    await dropbox_save(USERS_PATH, data)
    _cache_delete("users")
    if deleted_user_ids:
        companies_data = await dropbox_get(COMPANIES_PATH)
        if companies_data and companies_data.get("companies"):
            changed = False
            for c in companies_data["companies"]:
                before = c.get("assigned_users", []) or []
                after = [uid for uid in before if uid not in deleted_user_ids]
                if len(after) != len(before):
                    c["assigned_users"] = after
                    c["updated_at"] = datetime.now().isoformat()
                    changed = True
            if changed:
                await dropbox_save(COMPANIES_PATH, companies_data)
                _cache_delete("companies")
    return {"ok": True}

@app.post("/api/auth/login")
async def login(request: Request):
    body = await request.json()
    email = body.get("email", "").strip()
    password = body.get("password", "")
    data = await dropbox_get(USERS_PATH)
    if data is None:
        return JSONResponse({"error": "ユーザーが見つかりません"}, status_code=401)
    user = next((u for u in data.get("users", []) if u.get("email") == email), None)
    if not user:
        return JSONResponse({"error": "ユーザーが見つかりません"}, status_code=401)
    individual_pw = user.get("individual_password", "")
    if individual_pw:
        if password != individual_pw:
            return JSONResponse({"error": "パスワードが違います"}, status_code=401)
    else:
        if password != data.get("password", ""):
            return JSONResponse({"error": "パスワードが違います"}, status_code=401)
    password_changed = bool(individual_pw)
    for u in data.get("users", []):
        if u.get("email") == email:
            u["last_login"] = date.today().isoformat()
            break
    await dropbox_save(USERS_PATH, data)
    return {**{k: v for k, v in user.items() if k != "individual_password"}, "password_changed": password_changed, "last_login": date.today().isoformat()}

@app.post("/api/auth/change_password")
async def change_password(request: Request):
    body = await request.json()
    email = body.get("email", "").strip()
    new_password = body.get("new_password", "")
    data = await dropbox_get(USERS_PATH)
    if data is None:
        return JSONResponse({"error": "ユーザーが見つかりません"}, status_code=404)
    for user in data.get("users", []):
        if user.get("email") == email:
            user["individual_password"] = new_password
            break
    else:
        return JSONResponse({"error": "ユーザーが見つかりません"}, status_code=404)
    await dropbox_save(USERS_PATH, data)
    _cache_delete("users")
    return {"ok": True}

@app.post("/api/auth/logout")
async def auth_logout(request: Request):
    body = await request.json()
    user_id = body.get("user_id")
    if not user_id:
        raise HTTPException(status_code=400)
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    for u in users_data.get("users", []):
        if u.get("id") == user_id:
            u["last_login"] = ""
            break
    await dropbox_save(USERS_PATH, users_data)
    _cache_delete("users")
    return {"ok": True}

@app.post("/api/auth/ping")
async def auth_ping(request: Request):
    body = await request.json()
    user_id = body.get("user_id")
    if not user_id:
        raise HTTPException(status_code=400)
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    today = date.today().isoformat()
    for u in users_data.get("users", []):
        if u.get("id") == user_id:
            u["last_login"] = today
            break
    await dropbox_save(USERS_PATH, users_data)
    _cache_delete("users")
    return {"ok": True}

def generate_invoice_pdf(inv: dict) -> io.BytesIO:
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
    font = 'HeiseiKakuGo-W5'
    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    c.setFont(font, 20)
    c.drawCentredString(width / 2, height - 60, "請　求　書")

    c.setFont(font, 10)
    c.drawString(380, height - 90, f"請求日：{inv.get('invoice_date', '')}")
    c.drawString(380, height - 105, f"支払期限：{inv.get('due_date', '')}")

    c.setFont(font, 12)
    c.drawString(40, height - 120, "税理士法人　士魂商才　御中")

    c.setFont(font, 10)
    c.drawString(350, height - 140, inv.get('user_name', ''))
    if inv.get('address'):
        c.drawString(350, height - 155, inv.get('address', ''))
    if inv.get('phone'):
        c.drawString(350, height - 170, f"TEL: {inv.get('phone', '')}")
    if inv.get('invoice_number'):
        c.drawString(350, height - 185, f"登録番号: {inv.get('invoice_number', '')}")

    c.setFont(font, 13)
    c.drawString(40, height - 210, f"ご請求金額（税込）：¥{inv.get('total', 0):,} 円")
    c.line(40, height - 218, 300, height - 218)

    y = height - 250
    c.setFont(font, 9)
    c.setFillColorRGB(1, 1, 1)
    c.rect(40, y - 2, 520, 16, fill=1)
    c.setFillColorRGB(0.118, 0.227, 0.373)
    c.rect(40, y - 2, 520, 16, fill=1)
    c.setFillColorRGB(1, 1, 1)
    c.drawString(44, y + 2, "品目")
    c.drawString(270, y + 2, "作業時間")
    c.drawString(340, y + 2, "単価（円/時）")
    c.drawString(450, y + 2, "金額（円）")
    c.setFillColorRGB(0, 0, 0)
    y -= 20

    for item in inv.get('items', []):
        c.setFont(font, 9)
        c.drawString(44, y, item.get('company', ''))
        c.drawString(270, y, item.get('hours_display', ''))
        c.drawString(340, y, f"{item.get('unit_price', 0):,}")
        c.drawString(450, y, f"{item.get('amount', 0):,}")
        c.line(40, y - 4, 560, y - 4)
        y -= 18

    for item in inv.get('special_items', []):
        c.drawString(44, y, item.get('content', ''))
        c.drawString(270, y, "—")
        c.drawString(340, y, "—")
        c.drawString(450, y, f"{item.get('amount', 0):,}")
        c.line(40, y - 4, 560, y - 4)
        y -= 18

    y -= 8
    subtotal = inv.get('subtotal', 0)
    tax = inv.get('tax', 0)
    total = inv.get('total', 0)
    c.setFont(font, 9)
    c.drawString(370, y, "小計")
    c.drawString(450, y, f"¥{subtotal:,}")
    y -= 16
    c.drawString(370, y, "消費税（10%）")
    c.drawString(450, y, f"¥{tax:,}")
    y -= 4
    c.line(360, y, 520, y)
    y -= 14
    c.setFont(font, 11)
    c.drawString(370, y, "合計（税込）")
    c.drawString(450, y, f"¥{total:,}")

    y -= 40
    c.setFont(font, 9)
    bank_parts = [inv.get('bank_name', ''), inv.get('bank_branch', ''),
                  f"（{inv.get('bank_type', '')}）", inv.get('bank_number', ''),
                  inv.get('bank_holder', '')]
    bank_str = "　".join(p for p in bank_parts if p and p != '（）')
    if bank_str:
        c.drawString(40, y, f"振込先：{bank_str}")

    c.save()
    buf.seek(0)
    return buf

@app.get("/api/invoices/zip/{year_month}")
async def download_invoices_zip(year_month: str, user_id: str = Query(None)):
    if not user_id:
        raise HTTPException(status_code=401)
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    current_user = next((u for u in users_data.get("users", []) if u.get("id") == user_id), None)
    if not current_user or current_user.get("role") not in ["admin", "soumu"]:
        raise HTTPException(status_code=403)
    invoices_data = await dropbox_get(get_invoices_path(year_month)) or {"invoices": []}
    target = [inv for inv in invoices_data.get("invoices", [])
              if inv.get("year_month") == year_month and inv.get("status") == "approved"]
    if not target:
        raise HTTPException(status_code=404, detail="承認済み請求書がありません")
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for inv in target:
            pdf_buf = generate_invoice_pdf(inv)
            safe_name = inv.get('user_name', inv.get('user_id', 'unknown'))
            zf.writestr(f"{safe_name}_{year_month}_請求書.pdf", pdf_buf.getvalue())
    zip_buf.seek(0)
    from urllib.parse import quote
    zip_filename = quote(f"請求書一括_{year_month}.zip")
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{zip_filename}"}
    )

@app.get("/api/invoices/excel/{year_month}")
async def download_invoices_excel(year_month: str, user_id: str = Query(None)):
    if not user_id:
        raise HTTPException(status_code=401)
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    current_user = next((u for u in users_data.get("users", []) if u.get("id") == user_id), None)
    if not current_user or current_user.get("role", "staff") not in ["admin", "soumu"]:
        raise HTTPException(status_code=403)
    data = await dropbox_get(get_invoices_path(year_month)) or {"invoices": []}
    invoices = [inv for inv in data.get("invoices", [])
                if inv.get("year_month") == year_month and inv.get("status") == "approved"]

    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year_month}請求明細"

    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")
    row_fill   = PatternFill(fill_type="solid", fgColor="EFF6FF")
    right      = Alignment(horizontal="right")

    headers = ["スタッフ名", "会社名・業務内容", "作業時間", "金額（税込）"]
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for inv in invoices:
        first_row = row_idx
        for item in inv.get("items", []):
            ws.cell(row=row_idx, column=1, value=inv.get("user_name", ""))
            ws.cell(row=row_idx, column=2, value=item.get("company", ""))
            h_cell = ws.cell(row=row_idx, column=3, value=item.get("hours", 0))
            h_cell.number_format = "0.00"
            h_cell.alignment = right
            a_cell = ws.cell(row=row_idx, column=4, value=item.get("amount", 0))
            a_cell.number_format = "#,##0"
            a_cell.alignment = right
            row_idx += 1
        for item in inv.get("special_items", []):
            ws.cell(row=row_idx, column=1, value=inv.get("user_name", ""))
            ws.cell(row=row_idx, column=2, value=item.get("content", ""))
            ws.cell(row=row_idx, column=3, value=None)
            a_cell = ws.cell(row=row_idx, column=4, value=item.get("amount", 0))
            a_cell.number_format = "#,##0"
            a_cell.alignment = right
            row_idx += 1
        for r in range(first_row, row_idx):
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = row_fill
        ws.append([])
        row_idx += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 15

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    from urllib.parse import quote
    filename = quote(f"請求明細_{year_month}.xlsx")
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"})

@app.get("/api/invoices/my/{user_id}/{year_month}")
async def get_my_invoice(user_id: str, year_month: str):
    data = await dropbox_get(get_invoices_path(year_month)) or {"invoices": []}
    inv = next((i for i in data.get("invoices", [])
                if i.get("user_id") == user_id and i.get("year_month") == year_month), None)
    return inv or {}

@app.get("/api/invoices")
async def get_invoices(year_month: str = Query(None), user_id: str = Header(None)):
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    current_user = next((u for u in users_data.get("users", []) if u.get("id") == user_id), None)
    if not current_user:
        raise HTTPException(status_code=401)
    role = current_user.get("role", "staff")
    if role == "admin" or role == "soumu":
        target_ids = {u["id"] for u in users_data["users"]}
    elif role == "leader":
        my_group = current_user.get("group", "")
        target_ids = {u["id"] for u in users_data["users"] if u.get("group") == my_group}
    else:
        raise HTTPException(status_code=403)
    data = await dropbox_get(get_invoices_path(year_month or str(date.today().year) + "-01")) or {"invoices": []}
    invoices = [i for i in data.get("invoices", []) if i.get("user_id") in target_ids]
    if year_month:
        invoices = [i for i in invoices if i.get("year_month") == year_month]
    return {"invoices": invoices}

@app.post("/api/invoices/submit")
async def submit_invoice(request: Request):
    body = await request.json()
    user_id = body.get("user_id")
    if not user_id:
        raise HTTPException(status_code=401)
    year_month = body.get("year_month", "")
    invoices_path = get_invoices_path(year_month)
    data = await dropbox_get(invoices_path) or {"invoices": []}
    data["invoices"] = [i for i in data.get("invoices", [])
                        if not (i.get("user_id") == user_id and i.get("year_month") == year_month)]
    new_inv = {
        "id": f"inv_{int(time.time())}",
        "user_id": user_id,
        "user_name": body.get("user_name", ""),
        "year_month": year_month,
        "status": "pending",
        "submitted_at": date.today().isoformat(),
        "approved_at": "", "approved_by": "",
        "rejected_at": "", "rejected_by": "", "reject_reason": "",
        "invoice_date": body.get("invoice_date", ""),
        "due_date": body.get("due_date", ""),
        "subject": body.get("subject", ""),
        "address": body.get("address", ""),
        "phone": body.get("phone", ""),
        "bank_name": body.get("bank_name", ""),
        "bank_branch": body.get("bank_branch", ""),
        "bank_type": body.get("bank_type", ""),
        "bank_number": body.get("bank_number", ""),
        "bank_holder": body.get("bank_holder", ""),
        "invoice_number": body.get("invoice_number", ""),
        "items": body.get("items", []),
        "special_items": body.get("special_items", []),
        "subtotal": body.get("subtotal", 0),
        "tax": body.get("tax", 0),
        "total": body.get("total", 0),
    }
    data["invoices"].append(new_inv)
    await dropbox_save(invoices_path, data)
    return {"ok": True, "invoice": new_inv}

@app.post("/api/invoices/approve")
async def approve_invoice(request: Request):
    body = await request.json()
    invoice_id = body.get("invoice_id")
    approver_id = body.get("approver_id")
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    approver = next((u for u in users_data.get("users", []) if u.get("id") == approver_id), None)
    if not approver or approver.get("role", "staff") == "staff":
        raise HTTPException(status_code=403)
    current_year = str(date.today().year)
    for year in [current_year, str(int(current_year) - 1)]:
        invoices_path = get_invoices_path(f"{year}-01")
        data = await dropbox_get(invoices_path) or {"invoices": []}
        inv = next((i for i in data.get("invoices", []) if i.get("id") == invoice_id), None)
        if inv:
            inv["status"] = "approved"
            inv["approved_at"] = date.today().isoformat()
            inv["approved_by"] = approver.get("name", approver_id)
            await dropbox_save(invoices_path, data)
            return {"ok": True}
    raise HTTPException(status_code=404)

@app.post("/api/invoices/reject")
async def reject_invoice(request: Request):
    body = await request.json()
    invoice_id = body.get("invoice_id")
    rejector_id = body.get("rejector_id")
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    rejector = next((u for u in users_data.get("users", []) if u.get("id") == rejector_id), None)
    if not rejector or rejector.get("role", "staff") == "staff":
        raise HTTPException(status_code=403)
    current_year = str(date.today().year)
    for year in [current_year, str(int(current_year) - 1)]:
        invoices_path = get_invoices_path(f"{year}-01")
        data = await dropbox_get(invoices_path) or {"invoices": []}
        inv = next((i for i in data.get("invoices", []) if i.get("id") == invoice_id), None)
        if inv:
            inv["status"] = "rejected"
            inv["rejected_at"] = date.today().isoformat()
            inv["rejected_by"] = rejector.get("name", rejector_id)
            inv["reject_reason"] = body.get("reason", "")
            await dropbox_save(invoices_path, data)
            return {"ok": True}
    raise HTTPException(status_code=404)

@app.get("/api/invoice/{user_id}/{year_month}")
async def get_invoice_data(user_id: str, year_month: str):
    path = f"{REPORTS_BASE}/{user_id}_{year_month}.json"
    data = await dropbox_get(path) or {"entries": []}
    companies: dict = {}
    for e in data.get("entries", []):
        company = e.get("company_name") or "（会社名なし）"
        companies[company] = companies.get(company, 0) + e.get("hours", 0)
    def to_hhmm(h: float) -> str:
        hrs = int(h); mins = round((h - hrs) * 60)
        return f"{hrs}:{mins:02d}"
    items = [{"company": c, "hours": round(h * 100) / 100, "hours_display": to_hhmm(h)}
             for c, h in sorted(companies.items(), key=lambda x: -x[1])]
    total_hours = sum(companies.values())
    return {"items": items, "total_hours": round(total_hours * 100) / 100,
            "total_hours_display": to_hhmm(total_hours)}

@app.get("/api/pledges")
async def get_pledges(user_id: str = Header(None), year_month: str = Query(None)):
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    current_user = next((u for u in users_data.get("users", []) if u.get("id") == user_id), None)
    if not current_user:
        raise HTTPException(status_code=401)
    role = current_user.get("role", "staff")
    target_ym = year_month or f"{date.today().year}-{str(date.today().month).zfill(2)}"
    data = await dropbox_get(get_pledges_path(target_ym)) or {"pledges": []}
    pledges = data.get("pledges", [])
    if role == "staff":
        pledges = [p for p in pledges if p.get("user_id") == user_id]
    elif role == "leader":
        my_group = current_user.get("group", "")
        group_user_ids = {u["id"] for u in users_data.get("users", []) if u.get("group") == my_group}
        pledges = [p for p in pledges if p.get("user_id") in group_user_ids]
    return {"pledges": pledges}


@app.post("/api/pledges/submit")
async def submit_pledge(request: Request):
    body = await request.json()
    user_id = body.get("user_id")
    if not user_id:
        raise HTTPException(status_code=401)
    year_month = body.get("year_month", "")
    pledges_path = get_pledges_path(year_month)
    data = await dropbox_get(pledges_path) or {"pledges": []}
    data["pledges"] = [p for p in data.get("pledges", [])
                       if not (p.get("user_id") == user_id and p.get("year_month") == year_month)]
    new_pledge = {
        "id": f"pl_{int(time.time())}",
        "user_id": user_id,
        "user_name": body.get("user_name", ""),
        "year_month": year_month,
        "submitted_at": date.today().isoformat(),
        "checklist": body.get("checklist", []),
    }
    data["pledges"].append(new_pledge)
    await dropbox_save(pledges_path, data)
    return {"ok": True, "pledge": new_pledge}


def generate_pledge_pdf(pledge: dict) -> io.BytesIO:
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
    font = 'HeiseiKakuGo-W5'
    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    c.setFont(font, 16)
    c.drawCentredString(width / 2, height - 60, "テレワーク誓約書")

    c.setFont(font, 9)
    c.drawString(40, height - 90, f"税理士法人士魂・商才")
    c.drawString(380, height - 90, f"スタッフ氏名：{pledge.get('user_name', '')}")
    c.drawString(380, height - 105, f"対象月：{pledge.get('year_month', '')}")
    c.drawString(380, height - 120, f"提出日：{pledge.get('submitted_at', '')}")

    c.line(40, height - 130, width - 40, height - 130)

    c.setFont(font, 8)
    intro = "私は、税理士法人士魂・商才（以下「甲」という。）から業務の委託を受け、テレワークにより業務を遂行するにあたり、以下の事項を遵守することを誓約します。"
    c.drawString(40, height - 148, intro[:60])
    c.drawString(40, height - 160, intro[60:])

    articles = [
        "1. 契約及び法令の遵守",
        "2. 守秘義務・情報管理",
        "3. 情報セキュリティの確保",
        "4. 作業環境の適正化",
        "5. 資料・機器の管理・返却",
        "6. 再委託の禁止",
        "7. 事故等の報告義務",
        "8. 成果物の納期・品質",
        "9. 責任",
    ]
    y = height - 180
    for article in articles:
        c.drawString(50, y, article)
        y -= 14

    y -= 8
    c.setFont(font, 9)
    c.drawString(40, y, "【月次チェックリスト確認済み項目】")
    y -= 14
    c.setFont(font, 8)
    for item in pledge.get('checklist', []):
        c.drawString(50, y, f"✓ {item}")
        y -= 13

    y -= 10
    c.line(40, y, width - 40, y)
    y -= 16
    c.setFont(font, 9)
    c.drawString(40, y, "以上の内容を十分に理解し、誠実に遵守することを誓約します。")
    y -= 20
    c.drawString(300, y, "甲：税理士法人士魂・商才")
    y -= 14
    c.drawString(300, y, "代表社員　税理士　勝野　弘志　殿")
    y -= 20
    c.drawString(300, y, f"電子承認日：{pledge.get('submitted_at', '')}")

    c.save()
    buf.seek(0)
    return buf


@app.get("/api/pledges/zip/{year_month}")
async def download_pledges_zip(year_month: str, user_id: str = Query(None)):
    if not user_id:
        raise HTTPException(status_code=401)
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    current_user = next((u for u in users_data.get("users", []) if u.get("id") == user_id), None)
    if not current_user or current_user.get("role") not in ["admin", "soumu"]:
        raise HTTPException(status_code=403)
    data = await dropbox_get(get_pledges_path(year_month)) or {"pledges": []}
    targets = [p for p in data.get("pledges", []) if p.get("year_month") == year_month]
    if not targets:
        raise HTTPException(status_code=404, detail="提出済み誓約書がありません")
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for pledge in targets:
            pdf_buf = generate_pledge_pdf(pledge)
            safe_name = pledge.get('user_name', pledge.get('user_id', 'unknown'))
            zf.writestr(f"{safe_name}_{year_month}_誓約書.pdf", pdf_buf.getvalue())
    zip_buf.seek(0)
    from urllib.parse import quote
    zip_filename = quote(f"誓約書一括_{year_month}.zip")
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{zip_filename}"}
    )


@app.get("/api/pledges/my/{user_id}/{year_month}")
async def get_my_pledge(user_id: str, year_month: str):
    data = await dropbox_get(get_pledges_path(year_month)) or {"pledges": []}
    pledge = next((p for p in data.get("pledges", [])
                   if p.get("user_id") == user_id and p.get("year_month") == year_month), None)
    return pledge or {}


@app.post("/api/invoices/freee/{year_month}")
async def register_to_freee(year_month: str, request: Request):
    body = await request.json()
    user_id = body.get("user_id")

    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    current_user = next((u for u in users_data.get("users", []) if u.get("id") == user_id), None)

    if not current_user or current_user.get("role") not in ["admin", "soumu"]:
        raise HTTPException(status_code=403)

    invoices_data = await dropbox_get(get_invoices_path(year_month)) or {"invoices": []}

    targets = [
        inv for inv in invoices_data["invoices"]
        if inv["year_month"] == year_month
        and inv["status"] == "approved"
        and not inv.get("freee_deal_id")
    ]

    if not targets:
        return {"ok": True, "count": 0, "message": "登録対象がありません"}

    registered = 0
    errors = []

    async with httpx.AsyncClient(timeout=30) as client:
        for inv in targets:
            try:
                r = await client.post(
                    f"{SHIKONSHOSAI_APP_URL}/api/internal/register_deal",
                    json={
                        "secret":      INTERNAL_SECRET,
                        "issue_date":  inv.get("invoice_date", f"{year_month}-30"),
                        "due_date":    inv.get("due_date", ""),
                        "amount":      inv["total"],
                        "description": f"業務委託料 {year_month} {inv['user_name']}"
                    }
                )
                result = r.json()

                if result.get("ok"):
                    for i, item in enumerate(invoices_data["invoices"]):
                        if item["id"] == inv["id"]:
                            invoices_data["invoices"][i]["freee_deal_id"] = result["deal_id"]
                            invoices_data["invoices"][i]["freee_registered_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                            break
                    registered += 1
                else:
                    errors.append(f"{inv['user_name']}: {result.get('error', '不明なエラー')}")

            except Exception as e:
                errors.append(f"{inv['user_name']}: {str(e)}")

    await dropbox_save(get_invoices_path(year_month), invoices_data)

    return {
        "ok": True,
        "count": registered,
        "errors": errors,
        "message": f"{registered}件をfreeeに登録しました"
    }

@app.get("/api/ar/sheets")
async def get_ar_sheets(user_id: str = Header(None)):
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    current_user = next((u for u in users_data.get("users", []) if u.get("id") == user_id), None)
    if not current_user or current_user.get("role") not in ["admin", "leader", "soumu"]:
        raise HTTPException(status_code=403)
    try:
        ss = _get_spreadsheet()
        sheets = [ws.title for ws in ss.worksheets()]
        return {"ok": True, "sheets": sheets}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/ar/sheet/{sheet_name}")
async def get_ar_sheet(sheet_name: str, user_id: str = Header(None)):
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    current_user = next((u for u in users_data.get("users", []) if u.get("id") == user_id), None)
    if not current_user or current_user.get("role") not in ["admin", "leader", "soumu"]:
        raise HTTPException(status_code=403)
    try:
        ss = _get_spreadsheet()
        ws = ss.worksheet(sheet_name)
        rows = ws.get_all_values()
        return {"ok": True, "rows": rows}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ar/month_start")
async def ar_month_start(request: Request, user_id: str = Header(None)):
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    current_user = next((u for u in users_data.get("users", []) if u.get("id") == user_id), None)
    if not current_user or current_user.get("role") not in ["admin", "soumu"]:
        raise HTTPException(status_code=403)
    body = await request.json()
    year_month = body.get("year_month")  # "202605"

    try:
        ss = _get_spreadsheet()
        master_ws = ss.worksheet("マスタ")
        master_rows = master_ws.get_all_values()

        sheet_name = f"一覧表{year_month}"
        year = int(year_month[:4])
        month = int(year_month[4:])

        # 前月シート名
        if month == 1:
            prev_month = f"一覧表{year-1}12"
        else:
            prev_month = f"一覧表{year_month[:4]}{str(month-1).zfill(2)}"

        # 既存シートチェック
        existing_titles = [ws.title for ws in ss.worksheets()]
        if sheet_name in existing_titles:
            return {"ok": False, "message": f"{sheet_name}はすでに存在します"}

        # マスタの右隣に新シートを作成して挿入
        all_sheets = ss.worksheets()
        master_index = next(i for i, ws in enumerate(all_sheets) if ws.title == "マスタ")
        new_ws = ss.add_worksheet(title=sheet_name, rows=210, cols=22)
        others = [ws for ws in ss.worksheets() if ws.title != sheet_name]
        new_order = others[:master_index + 1] + [new_ws] + others[master_index + 1:]
        ss.reorder_worksheets(new_order)

        # 1行目ヘッダー
        header = master_rows[0] if master_rows else []
        updates = []
        if header:
            updates.append({"range": "A1", "values": [header]})

        # 2行目：合計行
        sum_row = ["合計", ""]
        for col_letter in ["D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V"]:
            sum_row.append(f"=SUM({col_letter}3:{col_letter}200)")
        updates.append({"range": "A2", "values": [sum_row]})

        # データ行（3行目〜）
        data_rows = master_rows[2:] if len(master_rows) > 2 else []
        for i, row in enumerate(data_rows):
            row_num = i + 3
            if not any(row):
                continue
            new_row = [
                row[0] if len(row) > 0 else "",
                row[1] if len(row) > 1 else "",
                row[2] if len(row) > 2 else "",
                f"='{prev_month}'!T{row_num}",
                "",
                f"=D{row_num}-IF(E{row_num}=\"\",0,E{row_num})",
                row[6] if len(row) > 6 else "",
                row[7] if len(row) > 7 else "",
                row[8] if len(row) > 8 else "",
                row[9] if len(row) > 9 else "",
                row[10] if len(row) > 10 else "",
                row[11] if len(row) > 11 else "",
                row[12] if len(row) > 12 else "",
                row[13] if len(row) > 13 else "",
                row[14] if len(row) > 14 else "",
                row[15] if len(row) > 15 else "",
                row[16] if len(row) > 16 else "",
                "",
                f"=SUM(G{row_num}:Q{row_num})",
                f"=F{row_num}+S{row_num}",
                "",
                f"=IF(U{row_num}=\"\",\"\",T{row_num}-U{row_num})",
            ]
            updates.append({"range": f"A{row_num}", "values": [new_row]})

        new_ws.batch_update(updates)

        # freeeに月次顧問料を売上登録
        issue_date = f"{year}-{str(month).zfill(2)}-01"
        token = await _get_freee_token()
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

        registered = []
        errors = []

        async with httpx.AsyncClient() as client:
            for i, row in enumerate(data_rows):
                if not any(row):
                    continue
                amount = 0
                for col_idx in range(6, 17):
                    try:
                        val = row[col_idx] if len(row) > col_idx else ""
                        if val:
                            amount += int(str(val).replace(",", ""))
                    except Exception:
                        pass
                if amount == 0:
                    continue
                freee_partner_name = row[2] if len(row) > 2 else ""
                if not freee_partner_name:
                    continue

                # partner_id取得
                pr = await client.get(
                    "https://api.freee.co.jp/api/1/partners",
                    headers=headers,
                    params={"company_id": FREEE_COMPANY_ID, "keyword": freee_partner_name}
                )
                partners = pr.json().get("partners", [])
                partner_id = next((p["id"] for p in partners if p["name"] == freee_partner_name), None)

                # 取引先がなければ新規登録
                if not partner_id:
                    cr = await client.post(
                        "https://api.freee.co.jp/api/1/partners",
                        headers=headers,
                        json={"company_id": FREEE_COMPANY_ID, "name": freee_partner_name}
                    )
                    if cr.status_code == 201:
                        partner_id = cr.json()["partner"]["id"]
                    else:
                        errors.append(f"{freee_partner_name}: 取引先登録失敗")
                        continue

                # 売上取引登録
                dr = await client.post(
                    "https://api.freee.co.jp/api/1/deals",
                    headers=headers,
                    json={
                        "company_id": FREEE_COMPANY_ID,
                        "issue_date": issue_date,
                        "type": "income",
                        "partner_id": partner_id,
                        "details": [{
                            "account_item_name": "売上高",
                            "tax_code": 1,
                            "item_name": "顧問報酬",
                            "amount": amount,
                            "description": f"顧問料 {year_month[:4]}年{str(month).zfill(2)}月"
                        }]
                    }
                )
                if dr.status_code == 201:
                    registered.append(freee_partner_name)
                else:
                    errors.append(f"{freee_partner_name}: {dr.text}")

        return {
            "ok": True,
            "sheet": sheet_name,
            "registered": registered,
            "errors": errors,
            "message": f"シート作成完了。freee登録: {len(registered)}件"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ar/mid_month")
async def ar_mid_month(request: Request, user_id: str = Header(None)):
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    current_user = next((u for u in users_data.get("users", []) if u.get("id") == user_id), None)
    if not current_user or current_user.get("role") not in ["admin", "soumu"]:
        raise HTTPException(status_code=403)
    body = await request.json()
    year_month = body.get("year_month")  # "202605"

    year = int(year_month[:4])
    month = int(year_month[4:])
    last_day = calendar.monthrange(year, month)[1]
    end_date = f"{year}-{str(month).zfill(2)}-{last_day}"

    try:
        ss = _get_spreadsheet()
        sheet_name = f"一覧表{year_month}"
        ws = ss.worksheet(sheet_name)
        rows = ws.get_all_values()

        token = await _get_freee_token()
        auth_headers = {"Authorization": f"Bearer {token}"}

        async with httpx.AsyncClient() as client:
            # 未決済income取引取得（入金額集計用）
            deals_resp = await client.get(
                "https://api.freee.co.jp/api/1/deals",
                headers=auth_headers,
                params={
                    "company_id": FREEE_COMPANY_ID,
                    "type": "income",
                    "status": "unsettled",
                    "issue_date_end": end_date,
                    "limit": 100
                }
            )
            income_by_partner: dict = {}
            for deal in deals_resp.json().get("deals", []):
                pname = deal.get("partner_name", "")
                income_by_partner[pname] = income_by_partner.get(pname, 0) + deal.get("amount", 0)

            # trial_bs：売掛金残高（取引先別）
            bs_resp = await client.get(
                "https://api.freee.co.jp/api/1/reports/trial_bs",
                headers=auth_headers,
                params={
                    "company_id": FREEE_COMPANY_ID,
                    "start_date": "2025-08-01",
                    "end_date": end_date,
                    "breakdown_display_type": "partner",
                    "account_item_display_type": "account_item"
                }
            )
            ar_balance: dict = {}
            for item in bs_resp.json().get("trial_bs", []):
                if "売掛金" in item.get("account_item_name", ""):
                    for partner in item.get("partners", []):
                        ar_balance[partner.get("partner_display_name", partner.get("name", ""))] = partner.get("closing_balance", 0)

        # スプレッドシート更新
        diffs = []
        batch_updates = []

        for i, row in enumerate(rows[2:]):
            row_num = i + 3
            if not any(row):
                continue
            freee_name = row[2] if len(row) > 2 else ""
            partner_name = row[1] if len(row) > 1 else ""
            if not freee_name:
                continue

            income = income_by_partner.get(freee_name, "")
            balance = ar_balance.get(freee_name, "")

            if income != "":
                batch_updates.append({"range": f"E{row_num}", "values": [[income]]})
            if balance != "":
                batch_updates.append({"range": f"U{row_num}", "values": [[balance]]})

            # 差額チェック
            try:
                t_val = row[19] if len(row) > 19 else ""
                if t_val and balance != "":
                    diff = int(str(t_val).replace(",", "")) - int(balance)
                    if diff != 0:
                        diffs.append({
                            "name": partner_name,
                            "freee_name": freee_name,
                            "t_balance": t_val,
                            "u_balance": balance,
                            "diff": diff
                        })
            except Exception:
                pass

        if batch_updates:
            ws.batch_update(batch_updates)

        return {
            "ok": True,
            "updated": len(batch_updates),
            "diffs": diffs,
            "message": f"{len(batch_updates)}セルを更新しました"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/ar/sync/{year_month}")
async def ar_sync(year_month: str, user_id: str = Header(None)):
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    current_user = next((u for u in users_data.get("users", []) if u.get("id") == user_id), None)
    if not current_user or current_user.get("role") not in ["admin", "soumu"]:
        raise HTTPException(status_code=403)

    year = int(year_month[:4])
    month = int(year_month[4:])
    issue_date = f"{year}-{str(month).zfill(2)}-01"

    try:
        ss = _get_spreadsheet()
        ws = ss.worksheet(f"一覧表{year_month}")
        rows = ws.get_all_values()

        token = await _get_freee_token()
        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

        updated = []
        errors = []

        async with httpx.AsyncClient() as client:
            dr = await client.get(
                "https://api.freee.co.jp/api/1/deals",
                headers=headers,
                params={
                    "company_id": FREEE_COMPANY_ID,
                    "type": "income",
                    "issue_date_start": issue_date,
                    "issue_date_end": issue_date,
                    "limit": 100
                }
            )
            existing_deals = {d["partner_name"]: d for d in dr.json().get("deals", [])}

            for i, row in enumerate(rows[2:]):
                row_num = i + 3
                if not any(row):
                    continue
                freee_name = row[2] if len(row) > 2 else ""
                if not freee_name:
                    continue

                amount = 0
                for col_idx in range(6, 17):
                    try:
                        val = row[col_idx] if len(row) > col_idx else ""
                        if val:
                            amount += int(str(val).replace(",", ""))
                    except Exception:
                        pass

                if amount == 0:
                    continue

                existing = existing_deals.get(freee_name)
                if existing and existing.get("amount") != amount:
                    pr = await client.patch(
                        f"https://api.freee.co.jp/api/1/deals/{existing['id']}",
                        headers=headers,
                        json={
                            "company_id": FREEE_COMPANY_ID,
                            "details": [{
                                "account_item_name": "売上高",
                                "tax_code": 1,
                                "item_name": "顧問報酬",
                                "amount": amount
                            }]
                        }
                    )
                    if pr.status_code == 200:
                        updated.append(freee_name)
                    else:
                        errors.append(f"{freee_name}: {pr.text}")

        return {
            "ok": True,
            "updated": updated,
            "errors": errors,
            "message": f"{len(updated)}件を更新しました"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


_companies_cache = None
_companies_cache_at = 0
COMPANIES_CACHE_TTL = 60 * 60 * 24  # 24時間

@app.get("/api/companies/upstream")
async def get_companies_upstream():
    global _companies_cache, _companies_cache_at

    if _companies_cache and time.time() - _companies_cache_at < COMPANIES_CACHE_TTL:
        return _companies_cache

    upstream_url = f"{SHIKONSHOSAI_APP_URL}/api/internal/companies_for_daily_report"
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(upstream_url, params={"secret": INTERNAL_SECRET})
    except Exception as e:
        print(f"[companies] 取得失敗 type={type(e).__name__} msg={e}")
        raise HTTPException(status_code=502, detail="upstream_unreachable")

    if r.status_code != 200:
        print(f"[companies] upstream非200 url={upstream_url} status={r.status_code} body={r.text[:500]!r}")
        raise HTTPException(status_code=502, detail="upstream_error")

    data = r.json()
    if isinstance(data, list) and len(data) > 0:
        _companies_cache = data
        _companies_cache_at = time.time()
    else:
        print(f"[companies] 上流が空配列または非list: type={type(data).__name__} len={len(data) if hasattr(data, '__len__') else 'N/A'}")
    return data


# ===== freee取引先同期ヘルパー =====

async def _sync_freee_partner(user: dict):
    """ユーザーの振込先情報をfreee取引先に同期する（失敗しても無視）"""
    if not SHIKONSHOSAI_APP_URL or not INTERNAL_SECRET:
        return
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            await client.post(
                f"{SHIKONSHOSAI_APP_URL}/api/internal/sync_partner",
                headers={"x-internal-secret": INTERNAL_SECRET},
                json={
                    "user_id": user.get("id") or user.get("email"),
                    "name": user.get("name", ""),
                    "invoice_number": user.get("invoice_number"),
                    "bank_name": user.get("bank_name"),
                    "bank_name_kana": user.get("bank_name_kana"),
                    "bank_branch": user.get("bank_branch"),
                    "branch_kana": user.get("branch_kana"),
                    "bank_account_type": user.get("bank_type"),
                    "bank_account_number": user.get("bank_number"),
                    "bank_account_name": user.get("bank_holder"),
                    "bank_code": user.get("bank_code"),
                    "branch_code": user.get("branch_code"),
                },
            )
    except Exception:
        pass


@app.get("/api/forecast")
async def get_forecast(user_id: str = Header(None)):
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    current_user = next((u for u in users_data.get("users", []) if u.get("id") == user_id), None)
    if not current_user:
        raise HTTPException(status_code=401)
    if current_user.get("role") not in ["admin", "leader", "soumu"]:
        raise HTTPException(status_code=403)
    try:
        ss = _get_spreadsheet()
        master_ws = ss.worksheet("マスタ")
        master_rows = master_ws.get_all_values()

        from datetime import date
        today = date.today()
        if today.month >= 8:
            fiscal_start_year = today.year
        else:
            fiscal_start_year = today.year - 1

        def get_fiscal_months(start_year):
            months = []
            for i in range(12):
                m = 8 + i
                y = start_year
                if m > 12:
                    m -= 12
                    y += 1
                months.append(f"{y}{str(m).zfill(2)}")
            return months

        all_months = []
        _y = today.year
        _m = today.month
        for _ in range(12):
            all_months.append(f"{_y}{str(_m).zfill(2)}")
            _m += 1
            if _m > 12:
                _m = 1
                _y += 1

        period1 = all_months[:12]  # 全12ヶ月
        period2 = []               # 翌期なし

        monthly_kanmon = {m: {"kanmon":0,"kiccho":0,"rental":0,"kyuyo":0,"sonota":0} for m in all_months}
        monthly_kessan = {m: 0 for m in all_months}

        for row in master_rows[2:]:
            if not any(row):
                continue
            mgmt_no = row[0] if len(row) > 0 else ""
            if not mgmt_no or len(mgmt_no) < 3:
                continue

            def to_int(val):
                try:
                    return int(str(val).replace(",", "").replace(" ", "")) if val else 0
                except:
                    return 0

            kanmon     = to_int(row[3]  if len(row) > 3  else 0)   # D列：顧問料
            kiccho     = to_int(row[4]  if len(row) > 4  else 0)   # E列：記帳代行
            rental     = sum(to_int(row[i] if len(row) > i else 0) for i in range(5, 12))  # F〜L列：レンタル料
            kyuyo      = to_int(row[12] if len(row) > 12 else 0)   # M列：給与計算
            sonota     = to_int(row[13] if len(row) > 13 else 0)   # N列：その他
            kessan_tax = to_int(row[18] if len(row) > 18 else 0)   # S列：決算報酬（税込）

            for m in all_months:
                monthly_kanmon[m]["kanmon"] += kanmon
                monthly_kanmon[m]["kiccho"] += kiccho
                monthly_kanmon[m]["rental"] += rental
                monthly_kanmon[m]["kyuyo"]  += kyuyo
                monthly_kanmon[m]["sonota"] += sonota

            if kessan_tax > 0:
                try:
                    km = int(mgmt_no[:3])
                    if 1 <= km <= 12:
                        bm = km + 2
                        byo = 0
                        if bm > 12:
                            bm -= 12
                            byo = 1
                        bill_year = today.year + byo
                        if km < today.month - 2:
                            bill_year += 1
                        bym = f"{bill_year}{str(bm).zfill(2)}"
                        if bym in monthly_kessan:
                            monthly_kessan[bym] += kessan_tax
                except:
                    pass

        def build_period_data(months):
            result = []
            for m in months:
                k = monthly_kanmon.get(m, {})
                subtotal = sum(k.get(x, 0) for x in ["kanmon","kiccho","rental","kyuyo","sonota"])
                kessan   = monthly_kessan.get(m, 0)
                result.append({
                    "month":    m,
                    "kanmon":   k.get("kanmon", 0),
                    "kiccho":   k.get("kiccho", 0),
                    "rental":   k.get("rental", 0),
                    "kyuyo":    k.get("kyuyo",  0),
                    "sonota":   k.get("sonota", 0),
                    "subtotal": subtotal,
                    "kessan":   kessan,
                    "total":    subtotal + kessan
                })
            return result

        end_m = today.month - 1 if today.month > 1 else 12
        end_y = today.year + 1 if today.month > 1 else today.year
        return {
            "ok": True,
            "period1": {
                "label": f"{today.year}年{today.month}月〜{end_y}年{end_m}月",
                "data":  build_period_data(period1)
            },
            "period2": {
                "label": "",
                "data":  []
            }
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/forecast/excel")
async def forecast_excel(user_id: str = Header(None)):
    users_data = await dropbox_get(USERS_PATH)
    if not users_data:
        raise HTTPException(status_code=500)
    current_user = next((u for u in users_data.get("users", []) if u.get("id") == user_id), None)
    if not current_user:
        raise HTTPException(status_code=401)
    if current_user.get("role") not in ["admin", "leader", "soumu"]:
        raise HTTPException(status_code=403)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from datetime import date

        today = date.today()
        fiscal_start_year = today.year if today.month >= 8 else today.year - 1

        def get_fiscal_months(start_year):
            months = []
            for i in range(12):
                m = 8 + i
                y = start_year
                if m > 12:
                    m -= 12
                    y += 1
                months.append((y, m, f"{y}{str(m).zfill(2)}"))
            return months

        all_yms_raw = []
        _y = today.year
        _m = today.month
        for _ in range(12):
            all_yms_raw.append((_y, _m, f"{_y}{str(_m).zfill(2)}"))
            _m += 1
            if _m > 12:
                _m = 1
                _y += 1

        period1 = all_yms_raw[:12]  # 全12ヶ月
        period2 = []                # 翌期なし
        all_yms = [x[2] for x in period1]

        ss = _get_spreadsheet()
        master_ws = ss.worksheet("マスタ")
        master_rows = master_ws.get_all_values()

        monthly_kanmon = {m: {"kanmon":0,"kiccho":0,"rental":0,"kyuyo":0,"sonota":0} for m in all_yms}
        monthly_kessan = {m: 0 for m in all_yms}

        for row in master_rows[2:]:
            if not any(row): continue
            mgmt_no = row[0] if len(row) > 0 else ""
            if not mgmt_no or len(mgmt_no) < 3: continue
            def ti(v):
                try: return int(str(v).replace(",","").replace(" ","")) if v else 0
                except: return 0
            kanmon = ti(row[3] if len(row)>3 else 0)   # D列：顧問料
            kiccho = ti(row[4] if len(row)>4 else 0)   # E列：記帳代行
            rental = sum(ti(row[i] if len(row)>i else 0) for i in range(5,12))  # F〜L列：レンタル料
            kyuyo  = ti(row[12] if len(row)>12 else 0)  # M列：給与計算
            sonota = ti(row[13] if len(row)>13 else 0)  # N列：その他
            kessan_tax = ti(row[18] if len(row)>18 else 0)  # S列：決算報酬（税込）
            for m in all_yms:
                monthly_kanmon[m]["kanmon"] += kanmon
                monthly_kanmon[m]["kiccho"] += kiccho
                monthly_kanmon[m]["rental"] += rental
                monthly_kanmon[m]["kyuyo"]  += kyuyo
                monthly_kanmon[m]["sonota"] += sonota
            if kessan_tax > 0:
                try:
                    km = int(mgmt_no[:3])
                    if 1 <= km <= 12:
                        bm = km + 2
                        byo = 0
                        if bm > 12:
                            bm -= 12
                            byo = 1
                        bill_year = today.year + byo
                        if km < today.month - 2:
                            bill_year += 1
                        bym = f"{bill_year}{str(bm).zfill(2)}"
                        if bym in monthly_kessan:
                            monthly_kessan[bym] += kessan_tax
                except: pass

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "売上予測"

        h_fill = PatternFill("solid", fgColor="DBEAFE")
        t_fill = PatternFill("solid", fgColor="F1F5F9")
        g_fill = PatternFill("solid", fgColor="DBEAFE")
        s_fill = PatternFill("solid", fgColor="F8F9FA")
        sum_fill = PatternFill("solid", fgColor="E8F0FE")
        bold = Font(bold=True)
        right = Alignment(horizontal="right")
        center = Alignment(horizontal="center")

        end_m = today.month - 1 if today.month > 1 else 12
        end_y = today.year + 1 if today.month > 1 else today.year
        period_label = f"{today.year}年{today.month}月〜{end_y}年{end_m}月"

        ws.cell(1,1).value = "項目"
        ws.cell(1,1).font = bold
        ws.cell(1,1).fill = h_fill
        for i,(y,m,_) in enumerate(period1):
            c = ws.cell(1, 2+i)
            c.value = f"{y}年{m}月"; c.font = bold; c.fill = h_fill; c.alignment = center
        ws.cell(1,14).value = f"{period_label} 合計"
        ws.cell(1,14).font = bold; ws.cell(1,14).fill = sum_fill; ws.cell(1,14).alignment = center

        items = [
            ("顧問料収入",         None,        "header"),
            ("　顧問料",            "kanmon",    "data"),
            ("　記帳代行",          "kiccho",    "data"),
            ("　レンタル料",        "rental",    "data"),
            ("　給与計算",          "kyuyo",     "data"),
            ("　その他",            "sonota",    "data"),
            ("顧問料小計",          "subtotal",  "total"),
            ("決算報酬収入",        None,        "header"),
            ("　決算報酬（税抜）",  "kessan",    "data"),
            ("合計（税抜）",        "total",     "grand"),
        ]

        for ri,(label,key,kind) in enumerate(items):
            rn = ri + 2
            fill = s_fill if kind=="header" else (t_fill if kind=="total" else (g_fill if kind=="grand" else None))
            fnt  = bold if kind in ("header","total","grand") else None
            c = ws.cell(rn,1); c.value = label
            if fill: c.fill = fill
            if fnt:  c.font = fnt

            s1 = 0
            for i,(_,_,ym) in enumerate(period1):
                k = monthly_kanmon.get(ym,{})
                sub = sum(k.get(x,0) for x in ["kanmon","kiccho","rental","kyuyo","sonota"])
                vm = {"kanmon":k.get("kanmon",0),"kiccho":k.get("kiccho",0),"rental":k.get("rental",0),
                      "kyuyo":k.get("kyuyo",0),"sonota":k.get("sonota",0),"subtotal":sub,
                      "kessan":monthly_kessan.get(ym,0),"total":sub+monthly_kessan.get(ym,0)}
                val = vm.get(key,"") if key else ""
                cell = ws.cell(rn,2+i); cell.value = val if val!="" else None
                cell.alignment = right
                if fill: cell.fill = fill
                if fnt:  cell.font = fnt
                if val != "": s1 += val
            ws.cell(rn,14).value = s1 if key else ""
            ws.cell(rn,14).alignment = right; ws.cell(rn,14).font = bold; ws.cell(rn,14).fill = sum_fill

        ws.column_dimensions["A"].width = 20
        for col in range(2, 15):
            ws.column_dimensions[ws.cell(1,col).column_letter].width = 14

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        from fastapi.responses import StreamingResponse
        filename = f"売上予測_{today.year}{str(today.month).zfill(2)}-{end_y}{str(end_m).zfill(2)}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ========================================
# 会社カルテ機能
# ========================================

def _company_manual_path(company_id: str) -> str:
    return f"{COMPANY_MANUALS_BASE}/{company_id}.json"

def _company_schedule_path(company_id: str) -> str:
    return f"{COMPANY_SCHEDULES_BASE}/{company_id}.json"

def _now_iso() -> str:
    return datetime.now().isoformat()

# 会社カルテの業務属性フィールド（型ごとに分類）
_KARTE_ATTR_BOOL_FIELDS = (
    "tax_extension", "consumption_tax_extension",
    "withholding_tax", "withholding_special",
    "payroll", "mtg",
    "simplified_tax_check", "notification_filing",
    "reward_withholding",
)
_KARTE_ATTR_STR_FIELDS = ("entity_type", "consumption_tax", "payroll_closing_type", "payroll_payment_month")
_KARTE_ATTR_INT_FIELDS = ("payroll_day", "mtg_day", "payroll_closing_day", "payroll_payment_day")
_KARTE_ATTR_FIELDS = _KARTE_ATTR_BOOL_FIELDS + _KARTE_ATTR_STR_FIELDS + _KARTE_ATTR_INT_FIELDS

def _coerce_attr_value(field: str, value):
    if field in _KARTE_ATTR_BOOL_FIELDS:
        return bool(value)
    if field in _KARTE_ATTR_STR_FIELDS:
        return value if value else None
    if field in _KARTE_ATTR_INT_FIELDS:
        try:
            return int(value) if value not in (None, "", False) else None
        except (TypeError, ValueError):
            return None
    return value

def _csv_attr_value(field: str, raw: str):
    """CSVの文字列セルから属性値を生成。空欄は None / False。"""
    s = (raw or "").strip()
    if field in _KARTE_ATTR_BOOL_FIELDS:
        return s == "1"
    if field in _KARTE_ATTR_STR_FIELDS:
        return s or None
    if field in _KARTE_ATTR_INT_FIELDS:
        if not s:
            return None
        try:
            return int(s)
        except ValueError:
            return None
    return s

def _add_months(month: int, n: int) -> int:
    """1〜12 の範囲で月を加算"""
    return (month + n - 1) % 12 + 1

def _month_end_day(month: int) -> int:
    if month in (1, 3, 5, 7, 8, 10, 12):
        return 31
    if month in (4, 6, 9, 11):
        return 30
    return 28  # 2月（うるう年は別途扱わない）

def _generate_fixed_events(company: dict) -> list:
    """会社属性から固定イベントリストを生成する"""
    events = []
    fiscal_month = int(company.get("fiscal_month") or 0) or 3
    entity_type = company.get("entity_type") or "corporation"
    tax_extension = bool(company.get("tax_extension"))
    ct_extension = bool(company.get("consumption_tax_extension"))
    consumption_tax = company.get("consumption_tax") or "exempt"
    withholding = bool(company.get("withholding_tax"))
    withholding_special = bool(company.get("withholding_special"))
    payroll = bool(company.get("payroll"))
    payroll_day = int(company.get("payroll_day") or 25)

    def make_event(name, recurrence, **kwargs):
        ev = {"id": "fe" + uuid4().hex[:8], "name": name, "recurrence": recurrence, "notes": "", "auto_generated": True}
        ev.update(kwargs)
        return ev

    if entity_type == "corporation":
        tax_offset = 3 if tax_extension else 2
        tax_month = _add_months(fiscal_month, tax_offset)
        events.append(make_event("法人税申告期限", "yearly", month=tax_month, day=_month_end_day(tax_month)))

        if consumption_tax in ("standard", "simplified"):
            ct_offset = 5 if ct_extension else 2
            ct_month = _add_months(fiscal_month, ct_offset)
            events.append(make_event("消費税申告期限", "yearly", month=ct_month, day=_month_end_day(ct_month)))
            # 課税事業者は毎期判定が必要なため、決算月末に生成
            events.append(make_event("簡易課税判定", "yearly", month=fiscal_month, day=_month_end_day(fiscal_month)))

        # 事前確定給与届出：法人は全社必須（決算月+3ヶ月末）
        ps_month = _add_months(fiscal_month, 3)
        events.append(make_event("事前確定給与届出", "yearly", month=ps_month, day=_month_end_day(ps_month)))

        # 予定納税確認：申告期限月（決算月+2ヶ月、延長は無視）から3・6・9ヶ月後の月末
        tax_base_month = _add_months(fiscal_month, 2)
        for offset in (3, 6, 9):
            yotei_month = _add_months(tax_base_month, offset)
            events.append(make_event(
                "予定納税確認", "yearly",
                month=yotei_month, day=_month_end_day(yotei_month),
            ))

    elif entity_type == "individual":
        events.append(make_event("所得税申告期限", "yearly", month=3, day=15))
        if consumption_tax in ("standard", "simplified"):
            events.append(make_event("消費税申告期限", "yearly", month=3, day=31))
            events.append(make_event("簡易課税判定", "yearly", month=12, day=31))

    if withholding:
        if withholding_special:
            events.append(make_event("給与等源泉所得税（1〜6月分）", "yearly", month=7, day=10))
            events.append(make_event("給与等源泉所得税（7〜12月分）", "yearly", month=1, day=20))
        else:
            events.append(make_event("給与等源泉所得税", "monthly", day_of_month=10))

    # 報酬等源泉所得税（毎月10日固定・納期特例なし）
    if bool(company.get("reward_withholding")):
        events.append(make_event("報酬等源泉所得税", "monthly", day_of_month=10))

    if payroll:
        payment_day = company.get("payroll_payment_day")
        try:
            payment_day = int(payment_day) if payment_day not in (None, "") else None
        except (TypeError, ValueError):
            payment_day = None
        if payment_day:
            check_day = payment_day - 3
            if check_day <= 0:
                # 月またぎ：支払日1〜3日の場合、前月末近辺に補正（28+check_day）
                check_day = 28 + check_day
            check_day = max(1, min(31, check_day))
            events.append(make_event(
                "給与計算チェック", "monthly", day_of_month=check_day,
                notes=f"支払日({payment_day}日)の3日前",
            ))
        else:
            # 支払日未設定の場合はデフォルト25日
            events.append(make_event("給与計算チェック", "monthly", day_of_month=25))

    return events

async def _check_karte_edit_permission(user_id: str, company: dict) -> bool:
    """admin/leaderまたは担当者であればTrueを返す"""
    if not user_id:
        return False
    users_data = await dropbox_get(USERS_PATH) or {"users": []}
    users = users_data.get("users", [])
    user = next((u for u in users if u.get("id") == user_id), None)
    if not user:
        return False
    if user.get("role") in ("admin", "leader"):
        return True
    if user_id in (company.get("assigned_users") or []):
        return True
    return False

async def _require_company_edit_permission(user_id: str, company_id: str) -> dict:
    """指定会社の編集権限を確認し、会社データを返す。なければ403/404。"""
    data = await dropbox_get(COMPANIES_PATH) or {"companies": []}
    company = next((c for c in data.get("companies", []) if c.get("id") == company_id), None)
    if not company:
        raise HTTPException(status_code=404, detail="company not found")
    if not await _check_karte_edit_permission(user_id, company):
        raise HTTPException(status_code=403, detail="編集権限がありません")
    return company

# ----- 会社マスタ -----

@app.get("/api/companies")
async def get_karte_companies():
    cached = _cache_get("companies")
    if cached is not None:
        return cached
    data = await dropbox_get(COMPANIES_PATH) or {"companies": []}
    if "companies" not in data:
        data = {"companies": []}
    _cache_set("companies", data, ttl=CACHE_TTL_COMPANIES)
    return data

@app.post("/api/companies")
async def create_karte_company(request: Request):
    body = await request.json()
    data = await dropbox_get(COMPANIES_PATH) or {"companies": []}
    now = _now_iso()
    new_company = {
        "id": "c" + uuid4().hex[:8],
        "name": body.get("name", ""),
        "code": body.get("code", ""),
        "type": body.get("type", "bookkeeping"),
        "fiscal_month": body.get("fiscal_month") or 0,
        "industry": body.get("industry", ""),
        "contract_types": body.get("contract_types", []),
        "freee_enabled": bool(body.get("freee_enabled", False)),
        "notes": body.get("notes", ""),
        "assigned_users": body.get("assigned_users", []),
        "created_at": now,
        "updated_at": now,
    }
    for f in _KARTE_ATTR_FIELDS:
        new_company[f] = _coerce_attr_value(f, body.get(f))
    data.setdefault("companies", []).append(new_company)
    await dropbox_save(COMPANIES_PATH, data)
    _cache_delete("companies")
    return new_company

@app.put("/api/companies/{company_id}")
async def update_karte_company(company_id: str, request: Request):
    body = await request.json()
    user_id = body.get("user_id", "")
    data = await dropbox_get(COMPANIES_PATH) or {"companies": []}
    for c in data.get("companies", []):
        if c.get("id") == company_id:
            if not await _check_karte_edit_permission(user_id, c):
                raise HTTPException(status_code=403, detail="編集権限がありません")
            for field in ["name", "code", "type", "fiscal_month", "industry",
                          "contract_types", "freee_enabled", "notes", "assigned_users"]:
                if field in body:
                    c[field] = body[field]
            for f in _KARTE_ATTR_FIELDS:
                if f in body:
                    c[f] = _coerce_attr_value(f, body.get(f))
            c["updated_at"] = _now_iso()
            await dropbox_save(COMPANIES_PATH, data)
            _cache_delete("companies")
            return c
    raise HTTPException(status_code=404, detail="company not found")

@app.delete("/api/companies/{company_id}")
async def delete_karte_company(company_id: str):
    data = await dropbox_get(COMPANIES_PATH) or {"companies": []}
    before = len(data.get("companies", []))
    data["companies"] = [c for c in data.get("companies", []) if c.get("id") != company_id]
    if len(data["companies"]) == before:
        raise HTTPException(status_code=404, detail="company not found")
    await dropbox_save(COMPANIES_PATH, data)
    _cache_delete("companies")
    await dropbox_delete(_company_manual_path(company_id))
    await dropbox_delete(_company_schedule_path(company_id))
    return {"ok": True}

@app.post("/api/companies/import_csv")
async def import_karte_companies_csv(file: UploadFile = File(...)):
    raw = await file.read()
    text = None
    for enc in ("cp932", "shift_jis", "utf-8-sig", "utf-8"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise HTTPException(status_code=400, detail="encoding error")
    data = await dropbox_get(COMPANIES_PATH) or {"companies": []}
    companies = data.setdefault("companies", [])
    added = 0
    updated = 0
    errors = []
    reader = csv.DictReader(io.StringIO(text))
    now = _now_iso()
    for i, row in enumerate(reader, start=2):
        try:
            name = (row.get("name") or "").strip()
            code = (row.get("code") or "").strip()
            if not name:
                errors.append(f"行{i}: name必須")
                continue
            type_val = (row.get("type") or "bookkeeping").strip() or "bookkeeping"
            fiscal_raw = (row.get("fiscal_month") or "").strip()
            fiscal_month = int(fiscal_raw) if fiscal_raw else 0
            industry = (row.get("industry") or "").strip()
            ct_raw = (row.get("contract_types") or "").strip()
            contract_types = [s for s in ct_raw.split(";") if s] if ct_raw else []
            freee_enabled = (row.get("freee_enabled") or "0").strip() == "1"
            notes = row.get("notes") or ""
            attr_payload = {f: _csv_attr_value(f, row.get(f)) for f in _KARTE_ATTR_FIELDS if f in row}
            existing = next((c for c in companies if code and c.get("code") == code), None)
            if existing:
                existing.update({
                    "name": name, "code": code, "type": type_val,
                    "fiscal_month": fiscal_month, "industry": industry,
                    "contract_types": contract_types, "freee_enabled": freee_enabled,
                    "notes": notes, "updated_at": now,
                })
                existing.update(attr_payload)
                updated += 1
            else:
                base = {
                    "id": "c" + uuid4().hex[:8],
                    "name": name, "code": code, "type": type_val,
                    "fiscal_month": fiscal_month, "industry": industry,
                    "contract_types": contract_types, "freee_enabled": freee_enabled,
                    "notes": notes, "assigned_users": [],
                    "created_at": now, "updated_at": now,
                }
                # 未指定の属性は None / False で埋めて後方互換
                for f in _KARTE_ATTR_FIELDS:
                    base.setdefault(f, _coerce_attr_value(f, attr_payload.get(f)))
                companies.append(base)
                added += 1
        except Exception as e:
            errors.append(f"行{i}: {e}")
    await dropbox_save(COMPANIES_PATH, data)
    _cache_delete("companies")
    return {"added": added, "updated": updated, "errors": errors}

@app.post("/api/companies/assign_csv")
async def assign_companies_csv(file: UploadFile = File(...)):
    raw = await file.read()
    text = None
    for enc in ("cp932", "shift_jis", "utf-8-sig", "utf-8"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise HTTPException(status_code=400, detail="encoding error")
    users_data = await dropbox_get(USERS_PATH) or {"users": []}
    email_to_id = {(u.get("email") or "").strip(): u.get("id") for u in users_data.get("users", []) if u.get("email") and u.get("id")}
    companies_data = await dropbox_get(COMPANIES_PATH) or {"companies": []}
    code_to_company = {c.get("code"): c for c in companies_data.get("companies", []) if c.get("code")}
    updated = 0
    skipped = 0
    errors = []
    reader = csv.DictReader(io.StringIO(text))
    now = _now_iso()
    for i, row in enumerate(reader, start=2):
        code = (row.get("code") or "").strip()
        emails_str = (row.get("emails") or "").strip()
        if not code or not emails_str:
            skipped += 1
            continue
        company = code_to_company.get(code)
        if not company:
            errors.append(f"行{i}: 会社CD '{code}' が見つかりません")
            skipped += 1
            continue
        emails = [e.strip() for e in emails_str.split(";") if e.strip()]
        user_ids = []
        for email in emails:
            uid = email_to_id.get(email)
            if uid:
                user_ids.append(uid)
            else:
                errors.append(f"行{i}: メールアドレス '{email}' のユーザーが見つかりません")
        company["assigned_users"] = user_ids
        company["updated_at"] = now
        updated += 1
    await dropbox_save(COMPANIES_PATH, companies_data)
    _cache_delete("companies")
    return {"updated": updated, "skipped": skipped, "errors": errors}

@app.post("/api/companies/clear_cache")
async def clear_karte_companies_cache():
    _cache_delete("companies")
    return {"ok": True}

# ----- 手順書 -----

@app.get("/api/company_manuals/{company_id}")
async def get_company_manual(company_id: str):
    data = await dropbox_get(_company_manual_path(company_id))
    if not data:
        return {"company_id": company_id, "content": ""}
    return data

@app.put("/api/company_manuals/{company_id}")
async def put_company_manual(company_id: str, request: Request):
    body = await request.json()
    user_id = body.get("user_id", "")
    await _require_company_edit_permission(user_id, company_id)
    data = {
        "company_id": company_id,
        "content": body.get("content", ""),
        "updated_by": user_id,
        "updated_at": _now_iso(),
    }
    await dropbox_save(_company_manual_path(company_id), data)
    return data

# ----- スケジュール・申し送り -----

def _empty_schedule(company_id: str) -> dict:
    return {"company_id": company_id, "fixed_events": [], "single_events": [], "memos": []}

async def _load_schedule(company_id: str) -> dict:
    data = await dropbox_get(_company_schedule_path(company_id))
    if not data:
        return _empty_schedule(company_id)
    data.setdefault("company_id", company_id)
    data.setdefault("fixed_events", [])
    data.setdefault("single_events", [])
    data.setdefault("memos", [])
    return data

@app.get("/api/company_schedules/{company_id}")
async def get_company_schedule(company_id: str):
    return await _load_schedule(company_id)

@app.post("/api/company_schedules/{company_id}/fixed_events")
async def add_fixed_event(company_id: str, request: Request):
    body = await request.json()
    await _require_company_edit_permission(body.get("user_id", ""), company_id)
    data = await _load_schedule(company_id)
    event = {
        "id": "fe" + uuid4().hex[:8],
        "name": body.get("name", ""),
        "recurrence": body.get("recurrence", "monthly"),
        "notes": body.get("notes", ""),
    }
    if event["recurrence"] == "monthly":
        event["day_of_month"] = int(body.get("day_of_month") or 1)
    else:
        event["month"] = int(body.get("month") or 1)
        event["day"] = int(body.get("day") or 1)
    data["fixed_events"].append(event)
    await dropbox_save(_company_schedule_path(company_id), data)
    return event

@app.post("/api/company_schedules/{company_id}/generate_events")
async def generate_company_events(company_id: str, request: Request):
    """会社属性から固定イベントを自動生成して追加（既存は残す）"""
    body = await request.json() if (await request.body()) else {}
    user_id = body.get("user_id", "")
    companies_data = await dropbox_get(COMPANIES_PATH) or {"companies": []}
    company = next((c for c in companies_data.get("companies", []) if c.get("id") == company_id), None)
    if not company:
        raise HTTPException(status_code=404, detail="company not found")
    if not await _check_karte_edit_permission(user_id, company):
        raise HTTPException(status_code=403, detail="編集権限がありません")
    if not company.get("entity_type"):
        raise HTTPException(status_code=400, detail="先に基本情報の属性を設定してください")
    schedule = await _load_schedule(company_id)
    # 既存の自動生成イベントを削除（手動追加分は残す）
    schedule["fixed_events"] = [e for e in schedule.get("fixed_events", []) if not e.get("auto_generated")]
    new_events = _generate_fixed_events(company)
    schedule["fixed_events"].extend(new_events)
    await dropbox_save(_company_schedule_path(company_id), schedule)
    return {"added": len(new_events), "events": new_events}

@app.put("/api/company_schedules/{company_id}/fixed_events/{event_id}")
async def update_fixed_event(company_id: str, event_id: str, request: Request):
    body = await request.json()
    await _require_company_edit_permission(body.get("user_id", ""), company_id)
    data = await _load_schedule(company_id)
    for ev in data["fixed_events"]:
        if ev.get("id") == event_id:
            ev["name"] = body.get("name", ev.get("name", ""))
            ev["recurrence"] = body.get("recurrence", ev.get("recurrence", "monthly"))
            ev["notes"] = body.get("notes", ev.get("notes", ""))
            if ev["recurrence"] == "monthly":
                ev["day_of_month"] = int(body.get("day_of_month") or ev.get("day_of_month") or 1)
                ev.pop("month", None)
                ev.pop("day", None)
            else:
                ev["month"] = int(body.get("month") or ev.get("month") or 1)
                ev["day"] = int(body.get("day") or ev.get("day") or 1)
                ev.pop("day_of_month", None)
            await dropbox_save(_company_schedule_path(company_id), data)
            return ev
    raise HTTPException(status_code=404, detail="event not found")

@app.delete("/api/company_schedules/{company_id}/fixed_events/{event_id}")
async def delete_fixed_event(company_id: str, event_id: str, request: Request):
    user_id = request.query_params.get("user_id", "")
    await _require_company_edit_permission(user_id, company_id)
    data = await _load_schedule(company_id)
    before = len(data["fixed_events"])
    data["fixed_events"] = [e for e in data["fixed_events"] if e.get("id") != event_id]
    if len(data["fixed_events"]) == before:
        raise HTTPException(status_code=404, detail="event not found")
    await dropbox_save(_company_schedule_path(company_id), data)
    return {"ok": True}

@app.post("/api/company_schedules/{company_id}/single_events")
async def add_single_event(company_id: str, request: Request):
    body = await request.json()
    await _require_company_edit_permission(body.get("user_id", ""), company_id)
    data = await _load_schedule(company_id)
    event = {
        "id": "se" + uuid4().hex[:8],
        "name": body.get("name", ""),
        "date": body.get("date", ""),
        "notes": body.get("notes", ""),
        "completed": False,
        "completed_by": None,
        "completed_at": None,
    }
    data["single_events"].append(event)
    await dropbox_save(_company_schedule_path(company_id), data)
    return event

@app.put("/api/company_schedules/{company_id}/single_events/{event_id}")
async def update_single_event(company_id: str, event_id: str, request: Request):
    body = await request.json()
    await _require_company_edit_permission(body.get("user_id", ""), company_id)
    data = await _load_schedule(company_id)
    for ev in data["single_events"]:
        if ev.get("id") == event_id:
            ev["name"] = body.get("name", ev.get("name", ""))
            ev["date"] = body.get("date", ev.get("date", ""))
            ev["notes"] = body.get("notes", ev.get("notes", ""))
            await dropbox_save(_company_schedule_path(company_id), data)
            return ev
    raise HTTPException(status_code=404, detail="event not found")

@app.delete("/api/company_schedules/{company_id}/single_events/{event_id}")
async def delete_single_event(company_id: str, event_id: str, request: Request):
    user_id = request.query_params.get("user_id", "")
    await _require_company_edit_permission(user_id, company_id)
    data = await _load_schedule(company_id)
    before = len(data["single_events"])
    data["single_events"] = [e for e in data["single_events"] if e.get("id") != event_id]
    if len(data["single_events"]) == before:
        raise HTTPException(status_code=404, detail="event not found")
    await dropbox_save(_company_schedule_path(company_id), data)
    return {"ok": True}

@app.post("/api/company_schedules/{company_id}/single_events/{event_id}/complete")
async def complete_single_event(company_id: str, event_id: str, request: Request):
    body = await request.json()
    user_id = body.get("user_id", "")
    data = await _load_schedule(company_id)
    for ev in data["single_events"]:
        if ev.get("id") == event_id:
            if ev.get("completed"):
                ev["completed"] = False
                ev["completed_by"] = None
                ev["completed_at"] = None
            else:
                ev["completed"] = True
                ev["completed_by"] = user_id
                ev["completed_at"] = _now_iso()
            await dropbox_save(_company_schedule_path(company_id), data)
            return ev
    raise HTTPException(status_code=404, detail="event not found")

@app.post("/api/company_schedules/{company_id}/memos")
async def add_memo(company_id: str, request: Request):
    body = await request.json()
    user_id = body.get("user_id", "")
    await _require_company_edit_permission(user_id, company_id)
    data = await _load_schedule(company_id)
    memo = {
        "id": "m" + uuid4().hex[:8],
        "text": body.get("text", ""),
        "created_by": user_id,
        "created_at": _now_iso(),
    }
    data["memos"].insert(0, memo)
    await dropbox_save(_company_schedule_path(company_id), data)
    return memo

@app.delete("/api/company_schedules/{company_id}/memos/{memo_id}")
async def delete_memo(company_id: str, memo_id: str, request: Request):
    user_id = request.query_params.get("user_id", "")
    await _require_company_edit_permission(user_id, company_id)
    data = await _load_schedule(company_id)
    before = len(data["memos"])
    data["memos"] = [m for m in data["memos"] if m.get("id") != memo_id]
    if len(data["memos"]) == before:
        raise HTTPException(status_code=404, detail="memo not found")
    await dropbox_save(_company_schedule_path(company_id), data)
    return {"ok": True}

# ----- ホーム用スケジュール -----

def _company_progress_path(company_id: str) -> str:
    return f"{COMPANY_PROGRESS_BASE}{company_id}.json"


def _normalize_completed_events(raw) -> dict:
    """completed_events を必ず {event_id: {"user_name": ...}} 形式にする。
    指示書⑨時点で list 形式で保存されていたデータとの後方互換のため。"""
    if isinstance(raw, dict):
        return {k: (v if isinstance(v, dict) else {"user_name": ""}) for k, v in raw.items()}
    if isinstance(raw, list):
        return {eid: {"user_name": ""} for eid in raw}
    return {}


async def _load_company_progress(company_id: str) -> dict:
    """会社の当月完了済みイベントマップ {event_id: {"user_name": ...}} を返す。
    year_month が当月でなければ空 dict を返す。"""
    current_ym = date.today().strftime('%Y-%m')
    data = await dropbox_get(_company_progress_path(company_id))
    if not data or data.get('year_month') != current_ym:
        return {}
    return _normalize_completed_events(data.get('completed_events'))


@app.get("/api/company_progress/{company_id}")
async def get_company_progress(company_id: str):
    """会社の当月進捗を取得する。当月でない場合は空データを返す。"""
    current_ym = date.today().strftime('%Y-%m')
    data = await dropbox_get(_company_progress_path(company_id))
    if not data or data.get('year_month') != current_ym:
        return {"year_month": current_ym, "completed_events": {}}
    return {
        "year_month": current_ym,
        "completed_events": _normalize_completed_events(data.get('completed_events')),
    }


@app.post("/api/company_progress/{company_id}/toggle")
async def toggle_company_progress(company_id: str, request: Request):
    """固定イベントの完了状態をトグルする。完了時の user_name を記録する。"""
    body = await request.json()
    event_id = body.get('event_id')
    user_name = body.get('user_name', '') or ''
    if not event_id:
        raise HTTPException(status_code=400, detail="event_id is required")

    current_ym = date.today().strftime('%Y-%m')
    data = await dropbox_get(_company_progress_path(company_id))
    if not data or data.get('year_month') != current_ym:
        data = {"year_month": current_ym, "completed_events": {}}

    completed = _normalize_completed_events(data.get('completed_events'))
    if event_id in completed:
        del completed[event_id]
    else:
        completed[event_id] = {"user_name": user_name}
    data['year_month'] = current_ym
    data['completed_events'] = completed
    await dropbox_save(_company_progress_path(company_id), data)
    return data


@app.get("/api/home/schedules")
async def get_home_schedules(user_id: str = Query(...)):
    companies_data = await dropbox_get(COMPANIES_PATH) or {"companies": []}
    today = date.today()
    end = today + timedelta(days=30)
    # 担当会社のIDを先に集めて progress を並列取得
    assigned_company_ids = [
        c.get("id") for c in companies_data.get("companies", [])
        if user_id in (c.get("assigned_users") or [])
    ]
    progress_results = await asyncio.gather(
        *[_load_company_progress(cid) for cid in assigned_company_ids]
    )
    progress_map = dict(zip(assigned_company_ids, progress_results))

    results = []
    for c in companies_data.get("companies", []):
        if user_id not in (c.get("assigned_users") or []):
            continue
        cid = c.get("id")
        cname = c.get("name", "")
        sched = await _load_schedule(cid)
        for ev in sched.get("single_events", []):
            d_str = ev.get("date", "")
            try:
                d = date.fromisoformat(d_str)
            except (ValueError, TypeError):
                continue
            if not (today <= d <= end):
                continue
            results.append({
                "date": d.isoformat(),
                "company_id": cid,
                "company_name": cname,
                "event_id": ev.get("id"),
                "event_type": "single",
                "name": ev.get("name", ""),
                "notes": ev.get("notes", ""),
                "completed": False,
                "completed_by": "",
            })
        for ev in sched.get("fixed_events", []):
            recurrence = ev.get("recurrence", "monthly")
            if recurrence == "monthly":
                dom = int(ev.get("day_of_month") or 0)
                if dom < 1 or dom > 31:
                    continue
                candidates = []
                for offset in (0, 1, 2):
                    y = today.year
                    m = today.month + offset
                    while m > 12:
                        m -= 12
                        y += 1
                    last_day = calendar.monthrange(y, m)[1]
                    day_use = min(dom, last_day)
                    try:
                        candidates.append(date(y, m, day_use))
                    except ValueError:
                        pass
                for d in candidates:
                    if today <= d <= end:
                        cmap = progress_map.get(cid, {})
                        eid = ev.get("id")
                        is_done = eid in cmap
                        results.append({
                            "date": d.isoformat(),
                            "company_id": cid,
                            "company_name": cname,
                            "event_id": eid,
                            "event_type": "fixed_monthly",
                            "name": ev.get("name", ""),
                            "notes": ev.get("notes", ""),
                            "completed": is_done,
                            "completed_by": cmap.get(eid, {}).get("user_name", "") if is_done else "",
                        })
            else:
                m = int(ev.get("month") or 0)
                d_ = int(ev.get("day") or 0)
                if m < 1 or m > 12 or d_ < 1 or d_ > 31:
                    continue
                for y in (today.year, today.year + 1):
                    last_day = calendar.monthrange(y, m)[1]
                    try:
                        d = date(y, m, min(d_, last_day))
                    except ValueError:
                        continue
                    if today <= d <= end:
                        cmap = progress_map.get(cid, {})
                        eid = ev.get("id")
                        is_done = eid in cmap
                        results.append({
                            "date": d.isoformat(),
                            "company_id": cid,
                            "company_name": cname,
                            "event_id": eid,
                            "event_type": "fixed_yearly",
                            "name": ev.get("name", ""),
                            "notes": ev.get("notes", ""),
                            "completed": is_done,
                            "completed_by": cmap.get(eid, {}).get("user_name", "") if is_done else "",
                        })
    results.sort(key=lambda r: (r["date"], r["company_name"]))
    return {"schedules": results}


@app.get("/")
async def root(response: Response):
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return FileResponse("static/index.html")

app.mount("/", StaticFiles(directory="static", html=True), name="static")
