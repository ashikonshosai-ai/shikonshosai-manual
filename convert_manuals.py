import json
import openpyxl
from pathlib import Path

EXCEL_FILE = "外部委託マニュアル（作成中）.xlsx"
OUTPUT_LOCAL_DROPBOX = Path.home() / "Library/CloudStorage/Dropbox/400000_CC/shikonshosai/manuals.json"

wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)

sheet_to_category = {
    "全般":               "全般・ルール",
    "入力業務について注意事項":    "全般・ルール",
    "新人ガイダンス":          "新人ガイダンス",
    "アシスタント業務":         "月次業務",
    "入力担当者業務":          "月次業務",
    "エビデンス収集":          "月次業務",
    "決算業務":             "決算業務",
    "決算の電子申告後業務":       "決算業務",
    "申告書チェックの修正方法":     "決算業務",
    "電子納税":             "決算業務",
    "年末調整":             "年末調整",
    "源泉徴収関連":           "年末調整",
    "年末調整準備資料":         "年末調整",
    "承認ルート":            "承認ルート",
    "新規関与の流れ":          "新規関与",
    "アシスタント用":          "アシスタント用",
    "動画マニュアル目次":        "動画マニュアル",
}

categories = {}

for sheet_name in wb.sheetnames:
    if sheet_name == "目次":
        continue
    cat_name = sheet_to_category.get(sheet_name, "その他")
    if cat_name not in categories:
        categories[cat_name] = {
            "id": f"cat_{len(categories)+1}",
            "name": cat_name,
            "manuals": []
        }

    ws = wb[sheet_name]
    lines = []
    for row in ws.iter_rows(values_only=True):
        parts = [str(c).strip() for c in row if c is not None and str(c).strip() and str(c).strip() != "None"]
        if parts:
            lines.append("　".join(parts))

    content = "\n".join(lines)

    categories[cat_name]["manuals"].append({
        "id": f"m_{sheet_name}",
        "title": sheet_name,
        "content": content,
        "video_url": "",
        "updated_at": "2026-04-20",
        "updated_by": "インポート"
    })

result = {"categories": list(categories.values())}

# ローカルファイルに保存
with open("manuals.json", "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print("manuals.json を生成しました")

# ローカルDropboxフォルダに直接書き込み（自動同期）
OUTPUT_LOCAL_DROPBOX.parent.mkdir(parents=True, exist_ok=True)
with open(OUTPUT_LOCAL_DROPBOX, "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print(f"Dropboxフォルダに書き込み完了: {OUTPUT_LOCAL_DROPBOX}")

print(f"\nカテゴリ数: {len(result['categories'])}")
for cat in result['categories']:
    print(f"  {cat['name']}: {len(cat['manuals'])}件")
